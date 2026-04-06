#!/usr/bin/env python3
"""
extractor_chicken.py — U.S. Chicken Industry Spread Tracker
============================================================
Builds / refreshes  chicken.db  with quarterly data for the dashboard.

DATA SOURCES
  • Broiler composite wholesale price  → USDA ERS API  (monthly)
  • Chicken parts weekly prices         → USDA AMS API  (weekly, AMS-3646)
  • SBM + Corn prices                   → USDA AMS API  (weekly cost inputs)
  • PPC U.S. Gross Margin               → baked-in from current Excel history;
                                          update manually in the HARD_PPC dict
                                          after each quarterly earnings release.

USAGE
  pip install requests openpyxl
  python extractor_chicken.py

OUTPUT
  chicken.db  (SQLite, ~50-100 KB)

SCHEMA  — table: quarterly
  quarter    TEXT PRIMARY KEY   e.g. "1Q17"
  year_q     INTEGER            sortable: 20171, 20172 …
  bw         REAL               Broiler wholesale cts/lb   (quarterly avg)
  breast     REAL               Breast B/S cts/lb          (quarterly avg)
  leg_qtrs   REAL               Leg Quarters cts/lb        (quarterly avg)
  wings      REAL               Wings cts/lb               (quarterly avg)
  tenders    REAL               Tenderloins cts/lb         (quarterly avg)
  sbm        REAL               SBM Illinois FOB $/ton     (quarterly avg)
  corn       REAL               Corn Central IL $/bu       (quarterly avg)
  fc_spot    REAL               2.9802*corn + 0.03851*sbm  (current quarter)
  fc_0q5     REAL               0.5*fc_spot + 0.5*fc_prior (0.5Q lag)
  fc_1q5     REAL               0.5*fc_prior + 0.5*fc_2prior (1.5Q lag)
  ppc_us_gm  REAL               PPC U.S. Gross Margin %    (decimal, NULL if N/A)
  ppc_cnl_gm REAL               PPC Consolidated GM %      (decimal, NULL if N/A)
  updated_at TEXT               ISO timestamp of last update

SCHEMA  — table: weekly
  report_date TEXT PRIMARY KEY  ISO date "YYYY-MM-DD" of the USDA report week
  bw          REAL              Broiler composite wholesale cts/lb
  breast      REAL              Breast B/S cts/lb
  leg_qtrs    REAL              Leg Quarters cts/lb
  wings       REAL              Wings cts/lb
  tenders     REAL              Tenderloins cts/lb
  sbm         REAL              SBM Illinois FOB $/ton
  corn        REAL              Corn Central IL $/bu
  updated_at  TEXT              ISO timestamp of last update
  Note: only stores weeks where at least one value was fetched.
        Dashboard still uses the quarterly table — this table is for
        historical raw-data audit / future drill-down features.
"""

import sqlite3, calendar, math, os, sys, json, time
from datetime import datetime, date
from typing import Optional

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests")

# ─── Configuration ─────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(__file__), "chicken.db")
TIMEOUT  = 30   # HTTP timeout seconds
RETRY    = 3    # API retry attempts

# Quarter range to maintain in DB
FIRST_QUARTER = ("1Q17", 2017, 1)
# Last expected quarter:  build up through current + 1
_now = datetime.now()
LAST_YEAR = _now.year
LAST_Q    = (_now.month - 1) // 3 + 1   # current calendar quarter

# ─── Hard-coded PPC data ────────────────────────────────────────────────────────
# Update this dict after each quarterly earnings release.
# Keys: "1Q17" … format.  Values: (us_gm, cnl_gm) as decimals (e.g. 0.1085)
# us_gm = None means use cnl_gm as fallback.
HARD_PPC = {
    # ── 2017 ──
    "1Q17": (0.10845,  0.10845),
    "2Q17": (0.17793,  0.17793),
    "3Q17": (0.19458,  0.19458),
    "4Q17": (0.10315,  0.10315),
    # ── 2018 ──
    "1Q18": (0.09682,  None),
    "2Q18": (0.08104,  None),
    "3Q18": (0.07047,  None),
    "4Q18": (0.02648,  None),
    # ── 2019 ──
    "1Q19": (0.09032,  None),
    "2Q19": (0.12864,  None),
    "3Q19": (0.09950,  None),
    "4Q19": (0.06542,  None),
    # ── 2020 ──
    "1Q20": (0.07172,  None),
    "2Q20": (0.04893,  None),
    "3Q20": (0.09672,  None),
    "4Q20": (0.04856,  None),
    # ── 2021 ──
    "1Q21": (0.06640,  None),
    "2Q21": (0.10690,  None),
    "3Q21": (0.11270,  None),
    "4Q21": (0.11450,  None),
    # ── 2022 ──
    "1Q22": (0.16350,  None),
    "2Q22": (0.18780,  None),
    "3Q22": (0.15700,  None),
    "4Q22": (0.00990,  None),
    # ── 2023 ──
    "1Q23": (0.01580,  None),
    "2Q23": (0.04660,  None),
    "3Q23": (0.06860,  None),
    "4Q23": (0.07490,  None),
    # ── 2024 ──
    "1Q24": (0.09200,  None),
    "2Q24": (0.16980,  None),
    "3Q24": (0.17770,  None),
    "4Q24": (0.14600,  None),
    # ── 2025 ──
    "1Q25": (0.14130,  None),
    "2Q25": (0.17350,  None),
    "3Q25": (0.17040,  None),
    "4Q25": (0.10540,  None),
    # ── 2026 (update as released) ──
    # "1Q26": (None, None),
}

# ─── Quarter helpers ─────────────────────────────────────────────────────────
def quarter_label(yr: int, q: int) -> str:
    return f"{q}Q{str(yr)[2:]}"

def qstart(yr: int, q: int) -> datetime:
    return datetime(yr, (q-1)*3+1, 1)

def qend(yr: int, q: int) -> datetime:
    m = q*3
    return datetime(yr, m, calendar.monthrange(yr, m)[1])

def all_quarters():
    """Yield (year, q, label) from FIRST_QUARTER through current quarter."""
    fy, fq = FIRST_QUARTER[1], FIRST_QUARTER[2]
    for yr in range(fy, LAST_YEAR+1):
        for q in range(1, 5):
            if (yr == fy and q < fq): continue
            if (yr == LAST_YEAR and q > LAST_Q): continue
            yield yr, q, quarter_label(yr, q)

# ─── USDA API Key ────────────────────────────────────────────────────────────
# Set via environment variable USDA_API_KEY (e.g. in GitHub Actions secrets).
# Without a key the MARS API may return 403 on cloud runners.
USDA_API_KEY = os.environ.get("USDA_API_KEY", "").strip()

# ─── HTTP helper ─────────────────────────────────────────────────────────────
def get_json(url: str, params: dict = None) -> dict:
    # USDA MARS API uses HTTP Basic Auth: api_key as username, empty password.
    # Do NOT send as query param — that returns 403 on cloud runners.
    auth = (USDA_API_KEY, "") if USDA_API_KEY else None
    for attempt in range(RETRY):
        try:
            r = requests.get(url, params=params, auth=auth, timeout=TIMEOUT)
            r.raise_for_status()
            data = r.json()
            # MARS API sometimes returns a bare list instead of {"results": [...]}
            if isinstance(data, list):
                return {"results": data}
            return data
        except Exception as e:
            if attempt == RETRY - 1:
                print(f"  ✗ HTTP error ({url[:60]}…): {e}")
                return {}
            time.sleep(2 ** attempt)
    return {}

# ─── USDA AMS weekly commodity fetcher ────────────────────────────────────────
# AMS MARS API: https://marsapi.ams.usda.gov/services/v1.2/reports/
# We use the "Livestock, Poultry, and Grain" report endpoints.
#
# AMS-3646  = National Chicken Parts   (weekly)
# AMS-3192  = Central Illinois Corn    (weekly)
# AMS-3511  = Soybean Meal             (weekly)
# USDA ERS  = Broiler Composite        (monthly, CSV download)

AMS_BASE = "https://marsapi.ams.usda.gov/services/v1.2/reports"

def fetch_ams_weekly(report_id: str, slug_filter: str, value_col: str,
                     date_from: str = "2016-01-01") -> list[dict]:
    """
    Fetch weekly AMS report data.
    Returns list of {'date': datetime, 'value': float}.
    """
    url = f"{AMS_BASE}/{report_id}"
    params = {
        "q":         slug_filter,
        "startDate": date_from,
        "endDate":   datetime.now().strftime("%m/%d/%Y"),
        "allSections": "true",
        "allCommodities": "true",
    }
    data = get_json(url, params)
    rows = []
    for item in data.get("results", []):
        try:
            dt = datetime.strptime(item.get("report_date",""), "%m/%d/%Y")
            v  = float(item.get(value_col, "") or "nan")
            if not math.isnan(v):
                rows.append({"date": dt, "value": v})
        except (ValueError, TypeError):
            continue
    return rows

def quarterly_avg(rows: list[dict], yr: int, q: int) -> Optional[float]:
    s, e = qstart(yr, q), qend(yr, q)
    vals = [r["value"] for r in rows if s <= r["date"] <= e]
    return sum(vals)/len(vals) if vals else None

# ─── Broiler Composite from USDA ERS ─────────────────────────────────────────
ERS_BW_URL = ("https://apps.fas.usda.gov/psdonline/circulars/livestock.pdf")
# Fallback: use AMS-2920 (Broiler Composite) or the ERS download
# The ERS table is not easily API-accessible; we use the USDA AMS
# Broiler report AMS-2020 (National Composite).

AMS_BROILER_ID = "2020"   # AMS Weekly National Composite Broiler report

def fetch_bw_wholesale() -> list[dict]:
    """
    Broiler composite wholesale cts/lb — PDF primary, MARS API fallback.
    """
    print("  Trying PDF source (pytbroilerfryer.pdf) …")
    rows = fetch_bw_from_pdf()
    if rows:
        return rows
    print("  PDF empty — falling back to MARS API …")
    url = f"{AMS_BASE}/{AMS_BROILER_ID}"
    params = {"startDate": "01/01/2016", "endDate": datetime.now().strftime("%m/%d/%Y"),
              "allSections": "true"}
    data = get_json(url, params)
    rows = []
    for item in data.get("results", []):
        try:
            dt = datetime.strptime(item.get("report_date", ""), "%m/%d/%Y")
            for field in ("weighted_avg", "price", "avg_price", "wtd_avg",
                          "composite_price", "national_composite"):
                v = item.get(field)
                if v is not None:
                    rows.append({"date": dt, "value": float(v)})
                    break
        except (ValueError, TypeError):
            continue
    return rows

# ─── Chicken Parts from USDA AMS Weekly PDF (ams_3646.pdf) ───────────────────
# URL: https://www.ams.usda.gov/mnreports/ams_3646.pdf  (overwritten weekly)
# Easier and more reliable than the MARS API endpoint.
# Parsed with pdfplumber; falls back gracefully if unavailable.

AMS_PARTS_PDF = "https://www.ams.usda.gov/mnreports/ams_3646.pdf"

def fetch_parts_from_pdf() -> dict:
    """
    Download the USDA AMS Weekly National Chicken Parts PDF and parse prices.
    Returns same format as fetch_parts(): dict of lists of {'date', 'value'}.
    Returns empty dict on any failure so caller can fall back to API.

    Strategy:
      1. Try pdfplumber table extraction (most reliable for tabular data)
      2. Fall back to regex on raw text if table extraction yields nothing
    """
    try:
        import pdfplumber, io, re
    except ImportError:
        print("  ⚠ pdfplumber not installed — skipping PDF fetch")
        return {}

    try:
        resp = requests.get(AMS_PARTS_PDF, timeout=TIMEOUT)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ✗ PDF download failed: {e}")
        return {}

    results = {"breast": [], "leg_qtrs": [], "wings": [], "tenders": []}

    try:
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            full_text  = "\n".join(page.extract_text() or "" for page in pdf.pages)
            all_tables = []
            for page in pdf.pages:
                tbls = page.extract_tables() or []
                all_tables.extend(tbls)

        # ── Extract report date ──────────────────────────────────────────────
        dt = None
        for pat in [
            r"[Ww]eek\s+[Oo]f\s+([A-Za-z]+ \d{1,2},?\s*\d{4})",
            r"[Ff]or\s+[Ww]eek\s+[Ee]nding\s+(\d{1,2}/\d{1,2}/\d{4})",
            r"(\d{2}/\d{2}/\d{4})",
        ]:
            m = re.search(pat, full_text)
            if m:
                raw = m.group(1).strip()
                for fmt in ("%B %d %Y", "%B %d, %Y", "%m/%d/%Y"):
                    try:
                        dt = datetime.strptime(raw, fmt); break
                    except ValueError:
                        continue
            if dt:
                break

        if not dt:
            print("  ⚠ PDF: could not parse report date")
            # Print first 800 chars of text to help diagnose format issues
            print("  PDF text preview:\n" + full_text[:800])
            return {}
        print(f"  PDF report date: {dt.strftime('%Y-%m-%d')}")
        # Debug: print lines containing key product words
        for line in full_text.splitlines():
            if any(kw in line.lower() for kw in ("breast","leg quarter","wing","tender")):
                print(f"  [PDF line] {line.strip()}")

        # ── Helper: grab weighted-average price from a text snippet ──────────
        def wtd_avg_from_snippet(snippet: str) -> float | None:
            # Pattern 1: price range followed by wtd avg  "NNN.NN-NNN.NN  NNN.NN"
            m = re.search(
                r"\d{2,3}\.\d{2}\s*[-–]\s*\d{2,3}\.\d{2}\s+([\d]{2,3}\.\d{2})",
                snippet)
            if m:
                return float(m.group(1))
            # Pattern 2: first decimal ≥ 10 in snippet
            nums = [float(n) for n in re.findall(r"\b(\d{2,3}\.\d{2})\b", snippet)
                    if float(n) >= 10]
            return nums[0] if nums else None

        # ── Method 1: scan extracted tables ──────────────────────────────────
        # Each table row is a list of cell strings. Look for rows whose first
        # cell matches a product keyword, then extract the price column.
        keyword_map = {
            "breast":   re.compile(r"breast.*b\.?/?s\.?|b\.?/?s\.?\s*breast|breast.*boneless",
                                   re.IGNORECASE),
            "leg_qtrs": re.compile(r"leg\s*quarter", re.IGNORECASE),
            "wings":    re.compile(r"wing", re.IGNORECASE),
            "tenders":  re.compile(r"tender|tenderloin", re.IGNORECASE),
        }
        found_via_table = {k: None for k in results}

        for table in all_tables:
            for row in table:
                cells = [str(c or "").strip() for c in row]
                row_text = " ".join(cells)
                for key, pat in keyword_map.items():
                    if found_via_table[key] is not None:
                        continue
                    if pat.search(row_text):
                        # Weighted avg is typically the 3rd or 4th numeric cell
                        nums = [float(c) for c in cells
                                if re.match(r"^\d{2,3}\.\d{2}$", c)]
                        if nums:
                            found_via_table[key] = nums[0]

        # ── Method 2: regex on raw text (fallback per field) ─────────────────
        text_patterns = {
            # "Breast - B/S:" — traço e abreviação B/S (não "Boneless Skinless")
            "breast":   r"[Bb]reast[,\s/\-]+B\.?/?S\.?|B\.?/?S\.?\s+[Bb]reast",
            "leg_qtrs": r"[Ll]eg\s+[Qq]uarters?",
            # "Wings - Whole:" — traço obrigatório evita falso match em "Previous Weeks"
            "wings":    r"[Ww]ings?\s*[-–]",
            "tenders":  r"[Tt]enderloins?|[Tt]enders?",
        }

        def find_via_text(pattern: str) -> float | None:
            m = re.search(pattern, full_text)
            if not m:
                return None
            return wtd_avg_from_snippet(full_text[m.end(): m.end() + 250])

        # Merge: prefer table result, then text result
        for key in results:
            price = found_via_table[key]
            if price is None:
                price = find_via_text(text_patterns[key])
            if price is not None:
                results[key].append({"date": dt, "value": price})
                print(f"  PDF {key}: {price:.2f} cts/lb")
            else:
                print(f"  ⚠ PDF {key}: not found in report")

    except Exception as e:
        import traceback
        print(f"  ✗ PDF parse error: {e}")
        traceback.print_exc()
        return {}

    return results


# ─── Chicken Parts from USDA AMS 3646 ────────────────────────────────────────
# AMS NW_LS644 = National Chicken Parts (weekly, cts/lb)
# Column mapping (may vary by API version):
#   Breast B/S   Leg Quarters   Wings   Tenderloins

def fetch_parts() -> dict[str, list[dict]]:
    """
    Returns dict with keys: 'breast', 'leg_qtrs', 'wings', 'tenders'
    Each: list of {'date': datetime, 'value': float}.
    Primary source: USDA AMS Weekly PDF (ams_3646.pdf) — no auth needed.
    Fallback: USDA AMS MARS API (report 3646) — requires API key.
    """
    # ── Primary: PDF ─────────────────────────────────────────────────────────
    print("  Trying PDF source (ams_3646.pdf) …")
    pdf_results = fetch_parts_from_pdf()
    has_pdf_data = any(len(v) > 0 for v in pdf_results.values())
    if has_pdf_data:
        return pdf_results
    print("  PDF empty or failed — falling back to MARS API …")

    # ── Fallback: MARS API ────────────────────────────────────────────────────
    url = f"{AMS_BASE}/3646"
    params = {
        "startDate": "09/01/2016",
        "endDate":   datetime.now().strftime("%m/%d/%Y"),
        "allSections": "true",
    }
    data = get_json(url, params)
    results = {k: [] for k in ("breast", "leg_qtrs", "wings", "tenders")}
    for item in data.get("results", []):
        try:
            dt = datetime.strptime(item.get("report_date",""), "%m/%d/%Y")
            # Map AMS fields → our keys  (field names vary; try several)
            field_map = {
                "breast":    ["breast_boneless_skinless","b_s_breast","breast","brest"],
                "leg_qtrs":  ["leg_quarters_bulk","leg_quarters","leg_qtrs","legquarters"],
                "wings":     ["wings_whole","wings","wing"],
                "tenders":   ["tenderloins","tenders","tenderloin"],
            }
            for key, field_candidates in field_map.items():
                for fc in field_candidates:
                    v = item.get(fc)
                    if v is not None:
                        try:
                            results[key].append({"date": dt, "value": float(v)})
                        except ValueError:
                            pass
                        break
        except (ValueError, TypeError):
            continue
    return results

# ─── Feed costs + BW from USDA PDF / TXT ─────────────────────────────────────
# URLs (overwritten weekly by USDA):
AMS_SBM_PDF  = "https://www.ams.usda.gov/mnreports/ams_3511.pdf"
AMS_CORN_TXT = "https://www.ams.usda.gov/mnreports/gx_gr115.txt"
AMS_BW_PDF   = "https://www.ams.usda.gov/mnreports/pytbroilerfryer.pdf"


def _parse_report_date(text: str) -> "datetime | None":
    """Extract report date from USDA report header text."""
    import re
    for pat in [
        r"[Ww]eek\s+[Oo]f\s+([A-Za-z]+ \d{1,2},?\s*\d{4})",
        r"[Ff]or\s+[Ww]eek\s+[Ee]nding[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
        r"[Dd]ate[:\s]+(\d{1,2}/\d{1,2}/\d{4})",
        r"\b(\d{1,2}/\d{1,2}/\d{4})\b",
        r"\b([A-Za-z]+ \d{1,2},\s*\d{4})\b",
    ]:
        m = re.search(pat, text)
        if m:
            raw = m.group(1).strip()
            for fmt in ("%B %d %Y", "%B %d, %Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(raw, fmt)
                except ValueError:
                    continue
    return None


def _first_price(text: str, min_val: float = 1.0, max_val: float = 9999.0) -> "float | None":
    """Return first decimal number in text within [min_val, max_val]."""
    import re
    for n in re.findall(r"\b(\d{1,4}\.\d{1,2})\b", text):
        v = float(n)
        if min_val <= v <= max_val:
            return v
    return None


def fetch_corn_from_txt() -> "list[dict]":
    """
    Corn Central Illinois $/bu from USDA AMS text report gx_gr115.txt.
    The report is daily/weekly and contains spot bid prices for Central IL.
    """
    import re
    try:
        resp = requests.get(AMS_CORN_TXT, timeout=TIMEOUT)
        resp.raise_for_status()
        text = resp.text
    except Exception as e:
        print(f"  ✗ Corn TXT download failed: {e}")
        return []

    dt = _parse_report_date(text)
    if not dt:
        print("  ⚠ Corn TXT: could not parse date")
        print("  TXT preview:\n" + text[:600])
        return []

    # Look for corn price: line containing "Corn" and a $/bu price (2-8 $/bu range)
    price = None
    for line in text.splitlines():
        if re.search(r"\bCorn\b", line, re.IGNORECASE):
            p = _first_price(line, min_val=2.0, max_val=15.0)
            if p:
                price = p
                print(f"  [TXT line] {line.strip()}")
                break

    if price is None:
        # Fallback: first price in 2-15 range in whole doc
        price = _first_price(text, min_val=2.0, max_val=15.0)

    if price:
        print(f"  Corn TXT {dt.strftime('%Y-%m-%d')}: {price:.2f} $/bu")
        return [{"date": dt, "value": price}]
    print("  ⚠ Corn TXT: price not found")
    return []


def fetch_sbm_from_pdf() -> "list[dict]":
    """
    SBM Illinois FOB Truck $/ton from USDA AMS PDF ams_3511.pdf.
    National Grain and Oilseed Processor Feedstuff Report (weekly).
    """
    import re
    try:
        import pdfplumber, io
    except ImportError:
        print("  ⚠ pdfplumber not installed — skipping SBM PDF")
        return []

    try:
        resp = requests.get(AMS_SBM_PDF, timeout=TIMEOUT)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ✗ SBM PDF download failed: {e}")
        return []

    try:
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        print(f"  ✗ SBM PDF parse error: {e}")
        return []

    dt = _parse_report_date(text)
    if not dt:
        print("  ⚠ SBM PDF: could not parse date")
        print("  PDF preview:\n" + text[:600])
        return []

    # Look for "Soybean Meal" + "Illinois" + "FOB Truck" → Wtd Avg price ($/ton, 200-700 range)
    price = None
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if re.search(r"[Ss]oybean\s+[Mm]eal", line):
            print(f"  [PDF line] {line.strip()}")
            # Search this line + next 3 lines for Illinois FOB price
            block = " ".join(lines[i:i+4])
            if re.search(r"[Ii]llinois|[Ii]L\b|FOB", block):
                p = _first_price(block, min_val=150.0, max_val=800.0)
                if p:
                    price = p
                    break
            # No Illinois qualifier found — try price directly on this line
            if price is None:
                p = _first_price(line, min_val=150.0, max_val=800.0)
                if p:
                    price = p

    if price:
        print(f"  SBM PDF {dt.strftime('%Y-%m-%d')}: {price:.2f} $/ton")
        return [{"date": dt, "value": price}]
    print("  ⚠ SBM PDF: price not found — printing relevant lines:")
    for line in lines:
        if re.search(r"[Ss]oybean|[Ss]BM|[Ii]llinois", line):
            print(f"    {line.strip()}")
    return []


def fetch_bw_from_pdf() -> "list[dict]":
    """
    Broiler composite wholesale cts/lb from USDA AMS pytbroilerfryer.pdf.
    """
    import re
    try:
        import pdfplumber, io
    except ImportError:
        print("  ⚠ pdfplumber not installed — skipping BW PDF")
        return []

    try:
        resp = requests.get(AMS_BW_PDF, timeout=TIMEOUT)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ✗ BW PDF download failed: {e}")
        return []

    try:
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        print(f"  ✗ BW PDF parse error: {e}")
        return []

    dt = _parse_report_date(text)
    if not dt:
        print("  ⚠ BW PDF: could not parse date")
        print("  PDF preview:\n" + text[:600])
        return []

    # Look for composite price line (cts/lb, 50-200 range)
    price = None
    lines = text.splitlines()
    for line in lines:
        if re.search(r"[Cc]omposite", line):
            print(f"  [PDF line] {line.strip()}")
            p = _first_price(line, min_val=40.0, max_val=250.0)
            if p:
                price = p
                break

    if price:
        print(f"  BW PDF {dt.strftime('%Y-%m-%d')}: {price:.2f} cts/lb")
        return [{"date": dt, "value": price}]
    print("  ⚠ BW PDF: price not found — printing all lines:")
    for line in lines[:40]:
        if line.strip():
            print(f"    {line.strip()}")
    return []


# ─── Feed costs from USDA AMS ────────────────────────────────────────────────
def fetch_sbm() -> list[dict]:
    """SBM Illinois FOB Truck $/ton — PDF primary, MARS API fallback."""
    print("  Trying PDF source (ams_3511.pdf) …")
    rows = fetch_sbm_from_pdf()
    if rows:
        return rows
    print("  PDF empty — falling back to MARS API …")
    url = f"{AMS_BASE}/3511"
    params = {"startDate": "01/01/2017", "endDate": datetime.now().strftime("%m/%d/%Y"),
              "allSections": "true"}
    data = get_json(url, params)
    rows = []
    for item in data.get("results", []):
        try:
            dt = datetime.strptime(item.get("report_date",""), "%m/%d/%Y")
            for field in ("illinois_fob_truck","il_fob_truck","price","avg_price"):
                v = item.get(field)
                if v is not None:
                    rows.append({"date": dt, "value": float(v)}); break
        except (ValueError, TypeError):
            continue
    return rows


def fetch_corn() -> list[dict]:
    """Corn Central Illinois $/bu — TXT primary, MARS API fallback."""
    print("  Trying TXT source (gx_gr115.txt) …")
    rows = fetch_corn_from_txt()
    if rows:
        return rows
    print("  TXT empty — falling back to MARS API …")
    url = f"{AMS_BASE}/3192"
    params = {"startDate": "01/01/2017", "endDate": datetime.now().strftime("%m/%d/%Y"),
              "allSections": "true"}
    data = get_json(url, params)
    rows = []
    for item in data.get("results", []):
        try:
            dt = datetime.strptime(item.get("report_date",""), "%m/%d/%Y")
            for field in ("central_illinois","central_il","price","avg_price"):
                v = item.get(field)
                if v is not None:
                    rows.append({"date": dt, "value": float(v)}); break
        except (ValueError, TypeError):
            continue
    return rows

# ─── Fallback: read from existing Excel files ─────────────────────────────────
def load_from_excel(base_dir: str) -> dict:
    """
    Fallback loader: reads from the IBBA Excel files if they exist.
    Returns dict of quarterly data keyed by label.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    data = {}  # keyed by (yr, q)

    # ─ Broiler Composite ─
    bw_path = os.path.join(base_dir, "Broiler Composite Price.xlsx")
    if os.path.exists(bw_path):
        wb = load_workbook(bw_path, data_only=True)
        ws = wb["broiler"]
        bw_rows = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            dt = r[0]
            if not dt or not isinstance(dt, datetime): continue
            v = r[1] if isinstance(r[1], (int, float)) else None
            if v: bw_rows.append({"date": dt, "value": v})
        print(f"  Excel BW: {len(bw_rows)} monthly rows")
        data["bw_rows"] = bw_rows

    # ─ Parts + Costs ─
    parts_path = os.path.join(base_dir, "US_Chicken_Weekly_Prices_IBBA.xlsx")
    if os.path.exists(parts_path):
        wb2 = load_workbook(parts_path, data_only=True)
        # Parts Weekly
        ws_p = wb2["Weekly Prices"]
        breast_rows, leg_rows, wings_rows, tenders_rows = [], [], [], []
        for r in ws_p.iter_rows(min_row=5, values_only=True):
            dt = r[0]
            if not dt or not isinstance(dt, datetime): continue
            if isinstance(r[1], (int,float)): breast_rows.append({"date":dt,"value":r[1]})
            if isinstance(r[2], (int,float)): leg_rows.append({"date":dt,"value":r[2]})
            if isinstance(r[3], (int,float)): wings_rows.append({"date":dt,"value":r[3]})
            if isinstance(r[4], (int,float)): tenders_rows.append({"date":dt,"value":r[4]})
        print(f"  Excel parts: breast={len(breast_rows)}, leg={len(leg_rows)}, "
              f"wings={len(wings_rows)}, tenders={len(tenders_rows)}")
        data["breast_rows"]  = breast_rows
        data["leg_rows"]     = leg_rows
        data["wings_rows"]   = wings_rows
        data["tenders_rows"] = tenders_rows
        # Cost Inputs
        ws_c = wb2["Cost Inputs"]
        sbm_rows, corn_rows = [], []
        for r in ws_c.iter_rows(min_row=3, values_only=True):
            dt = r[0]
            if not dt or not isinstance(dt, datetime): continue
            def sf(v):
                if v is None or v == '': return None
                try: return float(str(v).replace(';','.').replace(',',''))
                except: return None
            sbm_v  = sf(r[1])
            corn_v = sf(r[2])
            if sbm_v:  sbm_rows.append({"date":dt,"value":sbm_v})
            if corn_v: corn_rows.append({"date":dt,"value":corn_v})
        print(f"  Excel costs: sbm={len(sbm_rows)}, corn={len(corn_rows)}")
        data["sbm_rows"]  = sbm_rows
        data["corn_rows"] = corn_rows

    return data

# ─── Build weekly table ───────────────────────────────────────────────────────
def build_weekly_db(data: dict):
    """
    Upsert raw weekly observations into the `weekly` table.
    Existing rows are never deleted — only new weeks are inserted,
    and existing weeks are updated only if a NULL field gets a value.
    The quarterly table and dashboard are unaffected.
    """
    # Merge all date-keyed series into a single dict: date → {field: value}
    series_map = {
        "bw":       data.get("bw_rows",      []),
        "breast":   data.get("breast_rows",  []),
        "leg_qtrs": data.get("leg_rows",     []),
        "wings":    data.get("wings_rows",   []),
        "tenders":  data.get("tenders_rows", []),
        "sbm":      data.get("sbm_rows",     []),
        "corn":     data.get("corn_rows",    []),
    }

    weekly: dict[str, dict] = {}   # "YYYY-MM-DD" → {field: float}
    for field, rows in series_map.items():
        for r in rows:
            key = r["date"].strftime("%Y-%m-%d")
            weekly.setdefault(key, {})[field] = r["value"]

    if not weekly:
        print("  (no weekly rows to write)")
        return

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS weekly (
            report_date TEXT PRIMARY KEY,
            bw          REAL,
            breast      REAL,
            leg_qtrs    REAL,
            wings       REAL,
            tenders     REAL,
            sbm         REAL,
            corn        REAL,
            updated_at  TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_weekly_date ON weekly(report_date)")

    ts = datetime.now().isoformat(timespec="seconds")
    inserted = updated = 0

    for date_str, vals in sorted(weekly.items()):
        existing = cur.execute(
            "SELECT bw, breast, leg_qtrs, wings, tenders, sbm, corn FROM weekly WHERE report_date=?",
            (date_str,)
        ).fetchone()

        if existing is None:
            # New week — insert
            cur.execute("""
                INSERT INTO weekly (report_date, bw, breast, leg_qtrs, wings, tenders, sbm, corn, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                date_str,
                vals.get("bw"),     vals.get("breast"), vals.get("leg_qtrs"),
                vals.get("wings"),  vals.get("tenders"),
                vals.get("sbm"),    vals.get("corn"),
                ts
            ))
            inserted += 1
        else:
            # Existing week — only fill NULLs (never overwrite good data with NULL)
            fields = ["bw", "breast", "leg_qtrs", "wings", "tenders", "sbm", "corn"]
            updates = {}
            for i, f in enumerate(fields):
                if existing[i] is None and vals.get(f) is not None:
                    updates[f] = vals[f]
            if updates:
                set_clause = ", ".join(f"{f}=?" for f in updates)
                cur.execute(
                    f"UPDATE weekly SET {set_clause}, updated_at=? WHERE report_date=?",
                    (*updates.values(), ts, date_str)
                )
                updated += 1

    con.commit()
    n = cur.execute("SELECT COUNT(*) FROM weekly").fetchone()[0]
    latest = cur.execute(
        "SELECT report_date, bw, breast, sbm, corn FROM weekly ORDER BY report_date DESC LIMIT 3"
    ).fetchall()
    print(f"  Weekly table: {n} rows total  (+{inserted} new, {updated} updated)")
    print(f"  {'Date':<12} {'BW':>7} {'Breast':>8} {'SBM':>8} {'Corn':>7}")
    for r in latest:
        def f(v): return f"{v:7.2f}" if v is not None else "    N/A"
        print(f"  {r[0]:<12} {f(r[1])} {f(r[2]):>8} {f(r[3]):>8} {f(r[4]):>7}")
    con.close()


# ─── Load existing DB as baseline ─────────────────────────────────────────────
def load_db_baseline() -> dict:
    """
    Read existing chicken.db into a dict keyed by quarter label.
    Returns empty dict if DB does not exist or has no table.
    Used to preserve historical market data when APIs return empty results.
    """
    if not os.path.exists(DB_PATH):
        return {}
    try:
        con = sqlite3.connect(DB_PATH)
        con.row_factory = sqlite3.Row
        rows = con.execute("""
            SELECT quarter, bw, breast, leg_qtrs, wings, tenders,
                   sbm, corn, fc_spot, fc_0q5, fc_1q5,
                   ppc_us_gm, ppc_cnl_gm
            FROM quarterly
        """).fetchall()
        con.close()
        baseline = {}
        for r in rows:
            baseline[r["quarter"]] = dict(r)
        print(f"  ✓ Loaded {len(baseline)} rows from existing chicken.db as baseline")
        return baseline
    except Exception as e:
        print(f"  ⚠ Could not read baseline from existing DB: {e}")
        return {}


# ─── Build database ───────────────────────────────────────────────────────────
def build_db(data: dict, baseline: dict = None):
    """
    Populate chicken.db from the data dictionary.

    baseline: dict keyed by quarter label with existing DB values.
              When a market value is None from new data, the baseline value
              is preserved.  This ensures historical data from Excel is never
              overwritten with NULLs when running without Excel (e.g. CI/CD).
    """
    if baseline is None:
        baseline = {}

    print(f"\nWriting to {DB_PATH} …")
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS quarterly (
            quarter    TEXT PRIMARY KEY,
            year_q     INTEGER,
            bw         REAL,
            breast     REAL,
            leg_qtrs   REAL,
            wings      REAL,
            tenders    REAL,
            sbm        REAL,
            corn       REAL,
            fc_spot    REAL,
            fc_0q5     REAL,
            fc_1q5     REAL,
            ppc_us_gm  REAL,
            ppc_cnl_gm REAL,
            updated_at TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_year_q ON quarterly(year_q)")

    def pick(new_val, label: str, field: str):
        """Return new_val if non-None, otherwise fall back to baseline."""
        if new_val is not None:
            return new_val
        return (baseline.get(label) or {}).get(field)

    rows_data = {}   # label → dict of raw values (pre-lag)
    quarters_list = list(all_quarters())

    for yr, q, label in quarters_list:
        bw_avg     = quarterly_avg(data.get("bw_rows", []),      yr, q)
        breast_avg = quarterly_avg(data.get("breast_rows", []),  yr, q)
        leg_avg    = quarterly_avg(data.get("leg_rows", []),      yr, q)
        wings_avg  = quarterly_avg(data.get("wings_rows", []),   yr, q)
        tenders_avg= quarterly_avg(data.get("tenders_rows", []), yr, q)
        sbm_avg    = quarterly_avg(data.get("sbm_rows", []),     yr, q)
        corn_avg   = quarterly_avg(data.get("corn_rows", []),    yr, q)

        # ── Fall back to baseline for any NULL market values ──────────────
        bw_avg      = pick(bw_avg,      label, "bw")
        breast_avg  = pick(breast_avg,  label, "breast")
        leg_avg     = pick(leg_avg,     label, "leg_qtrs")
        wings_avg   = pick(wings_avg,   label, "wings")
        tenders_avg = pick(tenders_avg, label, "tenders")
        sbm_avg     = pick(sbm_avg,     label, "sbm")
        corn_avg    = pick(corn_avg,    label, "corn")

        fc = (2.9802*corn_avg + 0.03851*sbm_avg) if (corn_avg and sbm_avg) else None
        # Prefer freshly computed fc_spot; fall back to baseline if still None
        fc = pick(fc, label, "fc_spot")

        ppc_us, ppc_cnl = HARD_PPC.get(label, (None, None))

        rows_data[label] = {
            "yr": yr, "q": q,
            "bw": bw_avg, "breast": breast_avg, "leg_qtrs": leg_avg,
            "wings": wings_avg, "tenders": tenders_avg,
            "sbm": sbm_avg, "corn": corn_avg,
            "fc_spot": fc,
            "ppc_us_gm": ppc_us, "ppc_cnl_gm": ppc_cnl,
        }

    # Compute lag columns
    labels_ordered = [ql for _, _, ql in quarters_list]
    for i, label in enumerate(labels_ordered):
        rd = rows_data[label]
        fc_cur  = rd["fc_spot"]
        fc_prev = rows_data[labels_ordered[i-1]]["fc_spot"] if i >= 1 else None
        fc_p2   = rows_data[labels_ordered[i-2]]["fc_spot"] if i >= 2 else None

        fc_0q5 = 0.5*fc_cur + 0.5*fc_prev if (fc_cur and fc_prev) else None
        fc_1q5 = 0.5*fc_prev + 0.5*fc_p2  if (fc_prev and fc_p2)  else None

        # Fall back to baseline lag values if still None
        fc_0q5 = pick(fc_0q5, label, "fc_0q5")
        fc_1q5 = pick(fc_1q5, label, "fc_1q5")

        yr, q = rd["yr"], rd["q"]
        year_q = yr*10 + q
        ts = datetime.now().isoformat(timespec="seconds")

        cur.execute("""
            INSERT OR REPLACE INTO quarterly
            (quarter, year_q, bw, breast, leg_qtrs, wings, tenders,
             sbm, corn, fc_spot, fc_0q5, fc_1q5,
             ppc_us_gm, ppc_cnl_gm, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            label, year_q,
            rd["bw"], rd["breast"], rd["leg_qtrs"], rd["wings"], rd["tenders"],
            rd["sbm"], rd["corn"], rd["fc_spot"], fc_0q5, fc_1q5,
            rd["ppc_us_gm"], rd["ppc_cnl_gm"], ts
        ))

    con.commit()
    n = cur.execute("SELECT COUNT(*) FROM quarterly").fetchone()[0]
    # Print summary
    sample = cur.execute("""
        SELECT quarter, bw, breast, leg_qtrs, sbm, corn, fc_spot, ppc_us_gm
        FROM quarterly
        ORDER BY year_q DESC LIMIT 6
    """).fetchall()
    print(f"  {n} rows written. Latest 6:")
    print(f"  {'Q':<7} {'BW':>7} {'Breast':>7} {'Leg':>7} {'SBM':>8} {'Corn':>7} {'FC':>7} {'PPC_GM':>9}")
    for row in sample:
        def f(v): return f"{v:7.2f}" if v is not None else "   N/A "
        def fp(v): return f"{v*100:8.2f}%" if v is not None else "     N/A"
        print(f"  {row[0]:<7} {f(row[1])} {f(row[2])} {f(row[3])} {f(row[4])} {f(row[5])} {f(row[6])} {fp(row[7])}")
    con.close()
    print(f"\n✓ chicken.db ready  ({os.path.getsize(DB_PATH)//1024} KB)")

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("="*60)
    print("U.S. Chicken Spread Tracker — Database Extractor")
    print("="*60)

    # ── Step 0: Load existing DB as baseline ─────────────────────────────────
    # This ensures historical market data is preserved even when Excel files
    # are absent (e.g. GitHub Actions) and USDA APIs return empty/sparse data.
    print("\n[0/2] Loading existing DB baseline (if any) …")
    baseline = load_db_baseline()

    # ── Step 1: Fetch fresh market data ──────────────────────────────────────
    # Try Excel files first (preferred for historical depth and accuracy).
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_base = os.path.join(script_dir, "..", "..", "U.S. Chicken")  # relative to repo structure
    if not os.path.isdir(excel_base):
        excel_base = os.path.join(os.path.expanduser("~"), "OneDrive",
                                  "Documentos", "Claude", "U.S. Chicken")  # Windows OneDrive fallback

    data = {}
    if os.path.isdir(excel_base):
        print(f"\n[1/2] Loading from Excel files at {excel_base} …")
        data = load_from_excel(excel_base)
    else:
        print(f"\n[1/2] Excel source not found — fetching from USDA APIs …")
        print("  (Historical gaps will be filled from existing DB baseline)")
        print("  Fetching broiler wholesale …")
        bw_rows = fetch_bw_wholesale()
        print(f"  → {len(bw_rows)} rows")
        print("  Fetching chicken parts (AMS-3646) …")
        parts = fetch_parts()
        for k, v in parts.items():
            print(f"  → {k}: {len(v)} rows")
        print("  Fetching SBM (AMS-3511) …")
        sbm_rows = fetch_sbm()
        print(f"  → {len(sbm_rows)} rows")
        print("  Fetching Corn (AMS-3192) …")
        corn_rows = fetch_corn()
        print(f"  → {len(corn_rows)} rows")
        data = {
            "bw_rows":      bw_rows,
            "breast_rows":  parts.get("breast",  []),
            "leg_rows":     parts.get("leg_qtrs", []),
            "wings_rows":   parts.get("wings",    []),
            "tenders_rows": parts.get("tenders",  []),
            "sbm_rows":     sbm_rows,
            "corn_rows":    corn_rows,
        }

    # ── Step 2a: Upsert raw weekly data ──────────────────────────────────────
    print(f"\n[2/3] Storing raw weekly data …")
    build_weekly_db(data)

    # ── Step 2b: Build / refresh quarterly DB, merging with baseline ─────────
    print(f"\n[3/3] Building quarterly database …")
    build_db(data, baseline=baseline)

if __name__ == "__main__":
    main()
