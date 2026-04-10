#!/usr/bin/env python3
"""
extractor_bz.py — Brazil Beef Export Spread Tracker
=====================================================
Builds / refreshes  beef_bz.db  with monthly and weekly spread data.

DATA SOURCES
  • SECEX/MDIC monthly exports  → balanca.economia.gov.br annual CSVs
                                   NCM 0201 (fresh) + 0202 (frozen)
  • BCB PTAX BRL/USD daily FX   → BCB OLINDA API (primary) / BCB SGS (fallback)
  • CEPEA Boi Gordo R$/arroba   → local XLS file (seeding only)
                                   stored in DB, not re-fetched on weekly runs

USAGE
  pip install requests openpyxl
  python extractor_bz.py                          # incremental (SECEX + FX only)
  python extractor_bz.py --init                   # full seed (needs CEPEA XLS)
  python extractor_bz.py --init --cepea PATH.xls  # explicit CEPEA file path

OUTPUT
  beef_bz.db  (SQLite)

SCHEMA — table: monthly
  period       TEXT PRIMARY KEY   e.g. "2006-01"
  year         INTEGER
  month        INTEGER
  secex_usd_kg REAL               FOB USD / kg net weight
  fx           REAL               avg BCB PTAX BRL/USD for the month
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               avg CEPEA R$/arroba ÷ 15 for the month
  spread       REAL               secex_brl_kg / cepea_r_kg
  updated_at   TEXT               ISO timestamp

SCHEMA — table: weekly
  start_date   TEXT PRIMARY KEY   ISO date "YYYY-MM-DD"
  end_date     TEXT               ISO date "YYYY-MM-DD"
  secex_usd_kg REAL               from SECEX weekly cumulative data
  fx           REAL               avg BCB PTAX BRL/USD for the period
  secex_brl_kg REAL               secex_usd_kg × fx
  cepea_r_kg   REAL               avg CEPEA R$/arroba ÷ 15 for the period
  spread       REAL               secex_brl_kg / cepea_r_kg
  updated_at   TEXT               ISO timestamp
"""

import sqlite3, os, sys, time, warnings
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

try:
    import requests
    # Suppress InsecureRequestWarning for Brazilian government domains
    # (balanca.economia.gov.br uses ICP-Brasil cert not trusted by default on Linux)
    from urllib3.exceptions import InsecureRequestWarning
    warnings.filterwarnings("ignore", category=InsecureRequestWarning)
except ImportError:
    sys.exit("Missing: pip install requests")

# Domains that require SSL verification disabled (ICP-Brasil / SERPRO chain)
_NO_VERIFY_HOSTS = ("balanca.economia.gov.br", "olinda.bcb.gov.br", "api.bcb.gov.br")

# ── Paths ──────────────────────────────────────────────────────────────────────
DB_PATH    = Path(__file__).parent / "beef_bz.db"
TIMEOUT    = 30
RETRY      = 3
NCM_CODES  = {"0201", "0202"}
ANO_INI    = 2006   # earliest SECEX year to seed

# ── Weekly historical data (Aug 2022 – Mar 2026, 176 weeks) ───────────────────
# Format: (start_date, end_date, price_usd_kg)
# price_usd_kg = Revenue_000USD * 1000 / Volume_tons / 1000 = Revenue / Volume_kg
WEEKLY_SEED = [
    # Format: (start_date, end_date, price_usd_kg, vol_tons_mtd_cumulative)
    # vol_tons is MTD cumulative; materialise() de-accumulates to weekly incremental
    # ── 2022 ────────────────────────────────────────────────────────────────────
    ("2022-08-01","2022-08-05",6.25,39395.1),
    ("2022-08-08","2022-08-12",6.31,88806.1),
    ("2022-08-15","2022-08-19",6.19,128548.7),
    ("2022-08-22","2022-08-31",5.92,203230.5),
    ("2022-09-01","2022-09-09",6.06,63676.0),
    ("2022-09-12","2022-09-16",5.9,114065.2),
    ("2022-09-19","2022-09-23",5.96,155047.6),
    ("2022-09-26","2022-09-30",6.07,203023.8),
    ("2022-10-03","2022-10-07",5.96,55618.4),
    ("2022-10-10","2022-10-15",5.88,91159.1),
    ("2022-10-17","2022-10-21",5.84,142692.9),
    ("2022-10-24","2022-10-31",5.71,188557.9),
    ("2022-11-01","2022-11-14",5.35,68876.2),
    ("2022-11-14","2022-11-18",5.16,98389.0),
    ("2022-11-21","2022-11-30",5.1,148843.6),
    ("2022-12-01","2022-12-09",5.03,47776.6),
    ("2022-12-12","2022-12-16",4.92,81915.5),
    ("2022-12-19","2022-12-23",4.89,116634.2),
    ("2022-12-26","2022-12-30",4.93,152797.9),
    # ── 2023 ────────────────────────────────────────────────────────────────────
    ("2023-01-02","2023-01-06",4.88,40795.9),
    ("2023-01-09","2023-01-13",4.85,77824.6),
    ("2023-01-16","2023-01-20",4.83,107095.6),
    ("2023-01-23","2023-01-27",4.81,160191.1),
    ("2023-02-01","2023-02-10",4.82,47166.5),
    ("2023-02-13","2023-02-17",4.89,91816.6),
    ("2023-02-20","2023-02-28",4.85,126449.7),
    ("2023-03-01","2023-03-10",4.87,67427.1),
    ("2023-03-13","2023-03-17",4.85,89807.1),
    ("2023-03-20","2023-03-31",4.63,107470.6),
    ("2023-04-03","2023-04-07",4.55,22594.5),
    ("2023-04-10","2023-04-14",4.72,43245.6),
    ("2023-04-17","2023-04-21",4.83,72445.1),
    ("2023-04-24","2023-04-28",4.94,110339.9),
    ("2023-05-01","2023-05-05",5.06,42809.4),
    ("2023-05-08","2023-05-12",5.07,75017.4),
    ("2023-05-15","2023-05-19",5.12,113818.7),
    ("2023-05-22","2023-05-31",5.13,168509.6),
    ("2023-06-01","2023-06-09",5.14,70322.5),
    ("2023-06-12","2023-06-16",5.06,114541.5),
    ("2023-06-19","2023-06-23",4.99,154625.3),
    ("2023-06-26","2023-06-30",4.95,192741.6),
    ("2023-07-03","2023-07-07",4.86,37378.3),
    ("2023-07-10","2023-07-14",4.8,76629.1),
    ("2023-07-17","2023-07-21",4.61,117590.0),
    ("2023-07-24","2023-07-31",4.7,160795.4),
    ("2023-08-01","2023-08-04",4.53,41267.3),
    ("2023-08-07","2023-08-11",4.48,83343.8),
    ("2023-08-14","2023-08-18",4.53,124650.1),
    ("2023-08-21","2023-08-31",4.51,185364.9),
    ("2023-09-01","2023-09-08",4.49,74840.6),
    ("2023-09-11","2023-09-15",4.53,119984.0),
    ("2023-09-18","2023-09-22",4.6,158835.7),
    ("2023-09-25","2023-09-29",4.58,195071.7),
    ("2023-10-02","2023-10-06",4.59,38119.9),
    ("2023-10-09","2023-10-13",4.62,91285.7),
    ("2023-10-16","2023-10-20",4.55,133591.5),
    ("2023-10-23","2023-10-31",4.61,186203.9),
    ("2023-11-01","2023-11-10",4.6,73205.0),
    ("2023-11-13","2023-11-17",4.56,119027.5),
    ("2023-11-20","2023-11-30",4.61,187976.8),
    ("2023-12-01","2023-12-08",4.59,64877.5),
    ("2023-12-11","2023-12-15",4.54,97395.3),
    ("2023-12-18","2023-12-22",4.53,166130.4),
    ("2023-12-25","2023-12-29",4.52,208439.4),
    # ── 2024 ────────────────────────────────────────────────────────────────────
    ("2024-01-01","2024-01-05",4.52,49835.2),
    ("2024-01-08","2024-01-12",4.54,86833.4),
    ("2024-01-15","2024-01-19",4.47,123021.6),
    ("2024-01-22","2024-01-26",4.51,168103.2),
    ("2024-01-29","2024-01-31",4.7,181690.3),
    ("2024-02-01","2024-02-09",4.58,50220.1),
    ("2024-02-12","2024-02-23",4.52,143478.2),
    ("2024-02-26","2024-02-29",4.47,179119.6),
    ("2024-03-01","2024-03-08",4.5,50612.2),
    ("2024-03-11","2024-03-15",4.51,84673.8),
    ("2024-03-18","2024-03-22",4.56,139942.0),
    ("2024-03-25","2024-03-29",4.54,166327.6),
    ("2024-04-01","2024-04-05",4.48,54698.2),
    ("2024-04-08","2024-04-12",4.55,104326.6),
    ("2024-04-15","2024-04-19",4.56,155943.2),
    ("2024-04-22","2024-04-26",4.53,203839.7),
    ("2024-04-29","2024-04-30",4.62,208053.3),
    ("2024-05-01","2024-05-10",4.49,75405.9),
    ("2024-05-13","2024-05-31",4.51,211976.0),
    ("2024-06-03","2024-06-07",4.46,57944.2),
    ("2024-06-10","2024-06-14",4.45,97254.7),
    ("2024-06-17","2024-06-21",4.5,146293.5),
    ("2024-06-24","2024-06-28",4.46,192571.4),
    ("2024-07-01","2024-07-05",4.44,54204.0),
    ("2024-07-08","2024-07-12",4.39,109584.0),
    ("2024-07-15","2024-07-19",4.41,162524.5),
    ("2024-07-22","2024-07-26",4.42,215619.3),
    ("2024-07-29","2024-07-31",4.33,237267.1),
    ("2024-08-01","2024-08-09",4.42,71371.7),
    ("2024-08-12","2024-08-16",4.43,109718.8),
    ("2024-08-19","2024-08-23",4.51,164042.3),
    ("2024-08-26","2024-08-30",4.38,217458.7),
    ("2024-09-02","2024-09-06",4.41,70984.2),
    ("2024-09-09","2024-09-13",4.49,139185.3),
    ("2024-09-16","2024-09-20",4.57,185486.8),
    ("2024-09-23","2024-09-27",4.61,251755.9),
    ("2024-10-01","2024-10-04",4.6,39866.9),
    ("2024-10-07","2024-10-11",4.6,101660.9),
    ("2024-10-14","2024-10-18",4.61,176399.1),
    ("2024-10-21","2024-10-25",4.74,236196.9),
    ("2024-10-28","2024-10-31",4.81,270332.3),
    ("2024-11-04","2024-11-08",4.82,73522.6),
    ("2024-11-11","2024-11-15",4.84,137340.3),
    ("2024-11-18","2024-11-22",4.98,179991.3),
    ("2024-11-25","2024-11-29",4.87,228132.5),
    ("2024-12-02","2024-12-06",4.94,43033.6),
    ("2024-12-09","2024-12-13",4.88,89335.1),
    ("2024-12-16","2024-12-20",5.06,127290.3),
    ("2024-12-23","2024-12-31",4.95,202569.2),
    # ── 2025 ────────────────────────────────────────────────────────────────────
    ("2025-01-02","2025-01-10",5.06,66397.7),
    ("2025-01-13","2025-01-17",5.01,112731.7),
    ("2025-01-20","2025-01-24",5.03,143317.4),
    ("2025-01-27","2025-01-31",4.99,180473.7),
    ("2025-02-03","2025-02-07",4.96,47385.5),
    ("2025-02-10","2025-02-14",4.94,99848.6),
    ("2025-02-17","2025-02-21",4.9,153143.1),
    ("2025-02-24","2025-02-28",4.9,190457.8),
    ("2025-03-03","2025-03-07",4.89,60545.0),
    ("2025-03-10","2025-03-14",4.86,117480.6),
    ("2025-03-17","2025-03-21",4.91,163297.9),
    ("2025-03-24","2025-03-31",4.95,215427.2),
    ("2025-04-01","2025-04-04",4.95,37420.6),
    ("2025-04-07","2025-04-11",4.97,98194.2),
    ("2025-04-14","2025-04-17",5.04,159327.8),
    ("2025-04-21","2025-04-25",5.1,211548.1),
    ("2025-04-28","2025-04-30",5.09,241583.8),
    ("2025-05-02","2025-05-09",5.1,67165.3),
    ("2025-05-12","2025-05-16",5.13,123005.5),
    ("2025-05-19","2025-05-23",5.33,173804.1),
    ("2025-05-26","2025-05-30",5.29,218073.7),
    ("2025-06-02","2025-06-06",5.37,64225.3),
    ("2025-06-09","2025-06-13",5.46,117245.9),
    ("2025-06-16","2025-06-20",5.48,168837.9),
    ("2025-06-23","2025-06-30",5.49,241098.7),
    ("2025-07-01","2025-07-04",5.54,48715.5),
    ("2025-07-07","2025-07-11",5.53,104193.7),
    ("2025-07-14","2025-07-18",5.57,172709.5),
    ("2025-07-21","2025-07-25",5.54,243904.9),
    ("2025-07-28","2025-07-31",5.59,276879.0),
    ("2025-08-01","2025-08-08",5.56,80470.4),
    ("2025-08-11","2025-08-15",5.73,135785.1),
    ("2025-08-18","2025-08-22",5.55,212925.3),
    ("2025-08-25","2025-08-29",5.59,268562.6),
    ("2025-09-01","2025-09-05",5.56,78338.9),
    ("2025-09-08","2025-09-12",5.7,137274.5),
    ("2025-09-15","2025-09-19",5.64,209645.3),
    ("2025-09-22","2025-09-26",5.58,294706.7),
    ("2025-09-29","2025-09-30",5.68,314689.9),
    ("2025-10-01","2025-10-10",5.55,111919.9),
    ("2025-10-13","2025-10-17",5.45,201346.8),
    ("2025-10-20","2025-10-24",5.58,276493.4),
    ("2025-10-27","2025-10-31",5.62,320559.4),
    ("2025-11-03","2025-11-07",5.51,100536.4),
    ("2025-11-10","2025-11-14",5.56,163699.5),
    ("2025-11-17","2025-11-21",5.41,238219.7),
    ("2025-11-24","2025-11-28",5.56,318493.4),
    ("2025-12-01","2025-12-05",5.62,76721.3),
    ("2025-12-08","2025-12-12",5.59,143577.8),
    ("2025-12-15","2025-12-19",5.56,218356.4),
    ("2025-12-22","2025-12-31",5.65,304977.1),
    # ── 2026 ────────────────────────────────────────────────────────────────────
    ("2026-01-05","2026-01-09",5.53,89307.3),
    ("2026-01-12","2026-01-16",5.58,126254.1),
    ("2026-01-19","2026-01-23",5.65,183782.9),
    ("2026-01-26","2026-01-30",5.56,231821.3),
    ("2026-02-02","2026-02-06",5.62,68344.0),
    ("2026-02-09","2026-02-13",5.57,136800.0),
    ("2026-02-16","2026-02-20",5.66,192708.7),
    ("2026-02-23","2026-02-27",5.76,235889.7),
    ("2026-03-02","2026-03-06",5.69,59986.7),
    ("2026-03-09","2026-03-13",5.85,115678.5),
    ("2026-03-16","2026-03-20",5.82,167061.8),
    ("2026-03-23","2026-03-31",5.89,233951.5),
]


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════
def init_db(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS _fx_raw (
        dt   TEXT PRIMARY KEY,
        fx   REAL
    );
    CREATE TABLE IF NOT EXISTS _cepea_raw (
        dt        TEXT PRIMARY KEY,
        r_arroba  REAL,
        r_kg      REAL
    );
    CREATE TABLE IF NOT EXISTS _secex_raw (
        year         INTEGER,
        month        INTEGER,
        rev_000usd   REAL,
        vol_tons     REAL,
        price_usd_kg REAL,
        PRIMARY KEY (year, month)
    );
    CREATE TABLE IF NOT EXISTS _weekly_raw (
        start_date   TEXT PRIMARY KEY,
        end_date     TEXT,
        price_usd_kg REAL,
        vol_tons     REAL    -- MTD cumulative tons from SECEX weekly bulletin
    );
    CREATE TABLE IF NOT EXISTS monthly (
        period       TEXT PRIMARY KEY,
        year         INTEGER,
        month        INTEGER,
        secex_usd_kg REAL,
        fx           REAL,
        secex_brl_kg REAL,
        cepea_r_kg   REAL,
        spread       REAL,
        updated_at   TEXT
    );
    CREATE TABLE IF NOT EXISTS weekly (
        start_date   TEXT PRIMARY KEY,
        end_date     TEXT,
        secex_usd_kg REAL,
        fx           REAL,
        secex_brl_kg REAL,
        cepea_r_kg   REAL,
        spread       REAL,
        vol_tons     REAL,   -- incremental weekly tons (de-accumulated from MTD)
        updated_at   TEXT
    );
    """)
    conn.commit()
    # ── Migrate existing DBs that pre-date these columns ─────────────────────
    for tbl, col in [("_weekly_raw", "vol_tons"), ("weekly", "vol_tons")]:
        try:
            conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} REAL")
            conn.commit()
            print(f"  [DB] Migrated: added {tbl}.{col}")
        except Exception:
            pass  # column already exists


# ══════════════════════════════════════════════════════════════════════════════
# HTTP HELPER
# ══════════════════════════════════════════════════════════════════════════════
def get(url, **kwargs):
    hdrs = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
    # Brazilian government servers use ICP-Brasil certs not trusted on Linux by default
    ssl_verify = not any(h in url for h in _NO_VERIFY_HOSTS)
    for attempt in range(RETRY):
        try:
            r = requests.get(url, headers=hdrs, timeout=TIMEOUT,
                             verify=ssl_verify, **kwargs)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt == RETRY - 1:
                print(f"  ✗ {url[:70]}…: {e}")
                return None
            time.sleep(2 ** attempt)
    return None


# ══════════════════════════════════════════════════════════════════════════════
# BCB PTAX FX
# ══════════════════════════════════════════════════════════════════════════════
def fetch_fx(conn):
    """Download BCB PTAX BRL/USD daily rates into _fx_raw."""
    start = datetime(ANO_INI, 1, 1).strftime("%m-%d-%Y")
    end   = datetime.now().strftime("%m-%d-%Y")
    rows  = []

    # Method A: BCB OLINDA
    url = (
        "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
        f"CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
        f"?@dataInicial='{start}'&@dataFinalCotacao='{end}'"
        "&$format=json&$select=cotacaoCompra,dataHoraCotacao"
    )
    r = get(url)
    if r:
        for item in r.json().get("value", []):
            try:
                rows.append((item["dataHoraCotacao"][:10], float(item["cotacaoCompra"])))
            except Exception:
                pass
        print(f"  [FX] BCB OLINDA: {len(rows)} rows")

    # Method B: BCB SGS series 1
    if not rows:
        url2 = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.1/dados"
            f"?formato=json&dataInicial=01/01/{ANO_INI}&dataFinal="
            + datetime.now().strftime("%d/%m/%Y")
        )
        r2 = get(url2)
        if r2:
            for item in r2.json():
                try:
                    dt = datetime.strptime(item["data"], "%d/%m/%Y").strftime("%Y-%m-%d")
                    rows.append((dt, float(item["valor"])))
                except Exception:
                    pass
            print(f"  [FX] BCB SGS: {len(rows)} rows")

    if not rows:
        print("  [FX] All methods failed — FX not updated.")
        return 0

    conn.executemany("INSERT OR REPLACE INTO _fx_raw(dt,fx) VALUES(?,?)", rows)
    conn.commit()
    print(f"  [FX] {len(rows)} rows stored.")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# SECEX MONTHLY
# ══════════════════════════════════════════════════════════════════════════════
def fetch_secex(conn, years=None):
    """Download MDIC annual CSVs and upsert into _secex_raw."""
    from io import StringIO
    try:
        import pandas as pd
    except ImportError:
        sys.exit("Missing: pip install pandas openpyxl")

    if years is None:
        years = range(ANO_INI, datetime.now().year + 1)

    # The annual CSV uses 8-digit NCM codes (e.g. "02011000").
    # NCM_CODES contains 4-digit chapter codes ("0201", "0202"), so we
    # match by prefix — str[:4] — rather than exact equality.
    BASE = "https://balanca.economia.gov.br/balanca/bd/comexstat-bd/ncm/EXP_{year}.csv"
    total = 0
    for yr in years:
        r = get(BASE.format(year=yr))
        if not r:
            continue
        try:
            df = pd.read_csv(StringIO(r.text), sep=";", dtype=str, low_memory=False)
            # Normalise: strip whitespace and zero-pad to 8 chars
            df["CO_NCM"] = df["CO_NCM"].str.strip().str.zfill(8)
            # Keep rows whose 4-digit chapter prefix matches NCM_CODES
            df = df[df["CO_NCM"].str[:4].isin(NCM_CODES)].copy()
            if df.empty:
                print(f"  [SECEX] {yr}: no beef rows")
                continue
            df["CO_MES"]     = df["CO_MES"].astype(int)
            df["KG_LIQUIDO"] = df["KG_LIQUIDO"].astype(float)
            df["VL_FOB"]     = df["VL_FOB"].astype(float)
            grp = df.groupby("CO_MES").agg(
                vol_kg=("KG_LIQUIDO","sum"), rev_usd=("VL_FOB","sum")
            ).reset_index()
            rows = []
            for _, row in grp.iterrows():
                m    = int(row["CO_MES"])
                vol  = float(row["vol_kg"]) / 1000      # tons
                rev  = float(row["rev_usd"]) / 1000     # 000 USD
                p    = (rev * 1000 / (vol * 1000)) if vol > 0 else None  # USD/kg
                rows.append((yr, m, rev, vol, p))
            conn.executemany(
                "INSERT OR REPLACE INTO _secex_raw(year,month,rev_000usd,vol_tons,price_usd_kg)"
                " VALUES(?,?,?,?,?)", rows
            )
            conn.commit()
            total += len(rows)
            print(f"  [SECEX] {yr}: {len(rows)} months")
        except Exception as ex:
            print(f"  [SECEX] {yr}: {ex}")
    return total


# ══════════════════════════════════════════════════════════════════════════════
# CEPEA (local XLS — seeding only)
# ══════════════════════════════════════════════════════════════════════════════
def load_cepea_xls(conn, path):
    """Load CEPEA Boi Gordo R$/arroba from local XLS/XLSX file."""
    import subprocess, tempfile
    try:
        import pandas as pd
    except ImportError:
        sys.exit("Missing: pip install pandas openpyxl")

    p = Path(path)
    if not p.exists():
        print(f"  [CEPEA] Not found: {path}")
        return 0

    df = None
    # Try direct read (XLSX)
    try:
        df = pd.read_excel(p, header=None, engine="openpyxl")
    except Exception:
        pass

    # LibreOffice fallback (OLE2 .xls)
    if df is None:
        try:
            tmp = tempfile.mkdtemp()
            out = Path(tmp) / (p.stem + "_conv.xlsx")
            subprocess.run(
                ["libreoffice", "--headless", "--convert-to", "xlsx", str(p), "--outdir", tmp],
                capture_output=True, timeout=90
            )
            if out.exists():
                df = pd.read_excel(out, header=None, engine="openpyxl")
                print("  [CEPEA] LibreOffice conversion OK")
        except Exception as ex:
            print(f"  [CEPEA] LibreOffice failed: {ex}")

    if df is None:
        print("  [CEPEA] Could not read file.")
        return 0

    rows = []
    for _, row in df.iterrows():
        for c in range(len(row) - 1):
            cell = row.iloc[c]
            dt   = None
            if isinstance(cell, (datetime, date)):
                dt = cell.date() if isinstance(cell, datetime) else cell
            elif isinstance(cell, str):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        dt = datetime.strptime(cell.strip(), fmt).date()
                        break
                    except Exception:
                        pass
            if dt is None or dt.year < 2000:
                continue
            try:
                val = float(str(row.iloc[c + 1]).replace(",", "."))
                if 50 < val < 2000:
                    rows.append((str(dt), val, round(val / 15.0, 6)))
            except Exception:
                pass

    if not rows:
        print("  [CEPEA] No valid rows parsed.")
        return 0

    conn.executemany(
        "INSERT OR REPLACE INTO _cepea_raw(dt,r_arroba,r_kg) VALUES(?,?,?)", rows
    )
    conn.commit()
    print(f"  [CEPEA] {len(rows)} rows loaded from {p.name}")
    return len(rows)


# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY RAW SEED
# ══════════════════════════════════════════════════════════════════════════════
def seed_weekly_raw(conn):
    """Seed _weekly_raw from WEEKLY_SEED with smart conflict resolution:
      - INSERT new rows normally.
      - On conflict (row already exists):
          * Restore price_usd_kg from seed IF current value is NULL or clearly
            wrong (outside 2–20 USD/kg), without touching vol_tons live data.
          * Clear vol_tons if it looks like a bad value (< 100 t, which is
            impossible for a full week of Brazil beef exports).
    """
    conn.executemany(
        """
        INSERT INTO _weekly_raw(start_date, end_date, price_usd_kg, vol_tons)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(start_date) DO UPDATE SET
            -- Restore seed price only when existing price is NULL or out of range
            price_usd_kg = CASE
                WHEN _weekly_raw.price_usd_kg IS NULL THEN excluded.price_usd_kg
                WHEN _weekly_raw.price_usd_kg < 2.0   THEN excluded.price_usd_kg
                WHEN _weekly_raw.price_usd_kg > 20.0  THEN excluded.price_usd_kg
                ELSE _weekly_raw.price_usd_kg
            END,
            -- Clear vol_tons if it looks like a parsing artefact (< 100 t)
            vol_tons = CASE
                WHEN _weekly_raw.vol_tons IS NOT NULL
                     AND _weekly_raw.vol_tons < 100.0 THEN NULL
                ELSE _weekly_raw.vol_tons
            END
        """,
        WEEKLY_SEED
    )
    conn.commit()
    print(f"  [WEEKLY] {len(WEEKLY_SEED)} rows seeded/validated in _weekly_raw.")


# ══════════════════════════════════════════════════════════════════════════════
# FETCH WEEKLY SECEX BULLETIN (price + MTD volume)
# ══════════════════════════════════════════════════════════════════════════════
def fetch_weekly_bulletin(conn):
    """
    Fetch the SECEX/MDIC weekly bulletin Excel ("Produto por Atividade Econômica"
    — CUCI classification) from balanca.economia.gov.br and update _weekly_raw
    with the latest week's price_usd_kg and vol_tons (MTD cumulative).

    The bulletin page is a React SPA — links are rendered via JavaScript and
    won't appear in raw HTML. This function uses three discovery strategies:
      1. Search the raw HTML for any xlsx URL (href attrs AND JS strings)
      2. Try a set of guessed URL patterns based on known SECEX file naming
      3. As a last resort, try the SECEX API endpoint directly

    Excel structure (per current month):
      US$ Mil | US$ Mil/avg | Toneladas | Toneladas/avg | Preço (US$/Ton) | Var%

    We extract for "Carne bovina fresca, refrigerada ou congelada":
      - US$ Mil   → revenue (MTD, thousands USD)
      - Toneladas → volume (MTD, tons)

    De-accumulation (MTD → weekly) happens later in materialise().
    """
    import io, re as _re

    BASE = "https://balanca.economia.gov.br"
    PAGE_URL = f"{BASE}/balanca/pg_principal_bc/principais_resultados.html"

    from datetime import date as _date, timedelta as _td
    PT_MON = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
              7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
    today  = _date.today()
    yr     = today.year
    mo     = today.month
    # ISO week number (SECEX files often use 2-digit week)
    wk_num = today.isocalendar()[1]

    # ── Step 1: fetch HTML and search broadly for xlsx links ──────────────────
    # NOTE: we use requests directly (not get()) to avoid the duplicate-headers
    #       bug that occurs when get()'s hdrs= and **kwargs both carry 'headers'.
    xlsx_url = None
    KEYWORDS = ("cuci", "produto", "semana", "boletim", "isic", "ativ")
    try:
        import requests as _req
        html_r = _req.get(
            PAGE_URL,
            headers={"Accept": "text/html,application/xhtml+xml",
                     "User-Agent": "Mozilla/5.0"},
            timeout=30,
            verify=False,
        )
        page_text = html_r.text

        # Search 1a: href attributes containing .xls / .xlsx
        href_links = _re.findall(r'href=["\']([^"\']+\.xlsx?)["\']',
                                 page_text, _re.IGNORECASE)
        # Search 1b: any URL-like string with .xlsx in entire page source
        #            (catches JS bundle strings like "/path/file.xlsx")
        all_xlsx = _re.findall(r'["\']([^"\']*\.xlsx?)["\']',
                                page_text, _re.IGNORECASE)

        candidates = href_links + all_xlsx
        print(f"  [BULLETIN] Page fetched ({len(page_text):,} chars). "
              f"xlsx candidates found: {len(candidates)}")

        for lnk in candidates:
            if any(k in lnk.lower() for k in KEYWORDS):
                xlsx_url = lnk if lnk.startswith("http") else f"{BASE}{lnk}"
                print(f"  [BULLETIN] Found via page scrape: {xlsx_url}")
                break

        if not xlsx_url and candidates:
            print(f"  [BULLETIN] Candidates (no keyword match): {candidates[:8]}")

    except Exception as exc:
        print(f"  [BULLETIN] Page fetch error: {exc}")

    # ── Step 2: guessed URL patterns (SECEX naming conventions) ──────────────
    if xlsx_url is None:
        print("  [BULLETIN] Trying guessed URL patterns …")
        guesses = []
        for w in range(wk_num, wk_num - 3, -1):    # current week and 2 prior
            w = max(w, 1)
            for tpl in (
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/AtividadeEconomica_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/PRODUTO_EXP_SEMANA_{yr}_{w:02d}.xlsx",
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{mo:02d}.xlsx",
                f"/balanca/bd/boletim/CUCI_EXP_SEMANA_{yr}_{mo:02d}_{today.day:02d}.xlsx",
            ):
                guesses.append(BASE + tpl)

        for url in guesses:
            try:
                import requests as _req
                probe = _req.head(url, timeout=10, verify=False,
                                  headers={"User-Agent": "Mozilla/5.0"})
                if probe.status_code == 200:
                    xlsx_url = url
                    print(f"  [BULLETIN] Guessed URL found: {xlsx_url}")
                    break
                else:
                    print(f"  [BULLETIN]   {probe.status_code} {url.split('/')[-1]}")
            except Exception:
                pass

    if xlsx_url is None:
        print("  [BULLETIN] Could not locate Excel file. "
              "Set env var BULLETIN_XLSX_URL to override, e.g.:")
        print("  export BULLETIN_XLSX_URL='https://balanca.economia.gov.br/balanca/bd/boletim/CUCI_EXP_SEMANA_XXXX_YY.xlsx'")
        # Try env var override as last resort
        import os
        xlsx_url = os.environ.get("BULLETIN_XLSX_URL")
        if xlsx_url:
            print(f"  [BULLETIN] Using env override: {xlsx_url}")
        else:
            return 0

    # ── Step 3: download Excel ────────────────────────────────────────────────
    print(f"  [BULLETIN] Downloading: {xlsx_url}")
    r2 = get(xlsx_url)
    if r2 is None:
        return 0

    try:
        import openpyxl, zipfile as _zf, re as _re2
        raw_bytes = r2.content
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception as exc1:
            # openpyxl bug: Excel has broken drawing references in the zip.
            # Fix: rebuild zip without drawing files and strip drawing rels.
            if "drawing" in str(exc1).lower() or "no item named" in str(exc1).lower():
                print(f"  [BULLETIN] Drawing ref error — stripping drawings and retrying …")
                buf_fix = io.BytesIO()
                with _zf.ZipFile(io.BytesIO(raw_bytes), 'r') as zin:
                    with _zf.ZipFile(buf_fix, 'w', _zf.ZIP_DEFLATED) as zout:
                        for item in zin.infolist():
                            if 'drawing' in item.filename.lower():
                                continue          # drop drawing files
                            data = zin.read(item.filename)
                            if item.filename.endswith('.rels'):
                                # Remove <Relationship> entries pointing to drawings
                                data = _re2.sub(
                                    rb'<Relationship[^>]+/drawing[^>]+/?>',
                                    b'', data
                                )
                            zout.writestr(item, data)
                buf_fix.seek(0)
                wb = openpyxl.load_workbook(buf_fix, data_only=True)
            else:
                print(f"  [BULLETIN] Excel parse error: {exc1}")
                return 0
    except Exception as exc:
        print(f"  [BULLETIN] Excel parse error: {exc}")
        return 0

    # ── Step 4: find the right sheet and parse headers ────────────────────────
    # The workbook may have multiple sheets; try the active one first, then all
    sheets_to_try = [wb.active] + [wb[s] for s in wb.sheetnames if wb[s] != wb.active]

    # The Setores_Produtos.xlsx structure (confirmed from debug):
    #   Header row:  [Descrição | US$ Mil | US$ Mil/MédDiária | Toneladas |
    #                 Ton/MédDiária | Preço (US$/Tonelada) | Variação (%)]
    #   Each category has 2 data cols: [current period | previous period]
    #   "Current period" is always the FIRST column after each category header.
    #   There are NO per-column period sub-headers; the period label appears
    #   only in the "Variação" column as "MonAno - MonAno".
    #   → We find category header columns DIRECTLY and use them as data columns.

    def _parse_sheet(ws):
        """
        Return (vol_tons_mtd, price_usd_kg) for the current period, or (None, None).

        Strategy: locate the category header row by scanning for "Toneladas"
        and "Preço" headers. The current-period data is in the same column as
        the category header (first column of each metric pair).
        """
        ton_col   = None   # Excel column (1-based) for "Toneladas" header
        price_col = None   # Excel column (1-based) for "Preço (US$/Ton)" header
        usd_col   = None   # Excel column (1-based) for "US$ Mil" header (fallback)
        cat_header_row = None

        for row in ws.iter_rows(max_row=30):
            for cell in row:
                v = str(cell.value or "").strip().lower()
                col = cell.column
                if v.startswith("ton") and "media" not in v and ton_col is None:
                    ton_col = col
                    cat_header_row = cell.row
                if ("preço" in v or "preco" in v or "us$/ton" in v) and price_col is None:
                    price_col = col
                if "us$ mil" in v and "media" not in v and usd_col is None:
                    usd_col = col
            # Stop once we found at least Toneladas
            if ton_col is not None and cat_header_row is not None:
                # Look for Preço on the same row or within 2 rows
                if price_col is not None or usd_col is not None:
                    break

        if ton_col is None or cat_header_row is None:
            return None, None

        # ── Find "Carne bovina fresca" row ──────────────────────────────────────
        carne_row = None
        for row in ws.iter_rows(min_row=cat_header_row + 1):
            for cell in row[:3]:
                v = str(cell.value or "").lower()
                if "carne bovina" in v and "fresca" in v:
                    carne_row = cell.row
                    break
            if carne_row:
                break

        if carne_row is None:
            return None, None

        # ── Extract values ───────────────────────────────────────────────────────
        vol_tons = ws.cell(row=carne_row, column=ton_col).value

        if price_col is not None:
            # Preferred: read price directly from "Preço (US$/Tonelada)" column
            price_per_ton = ws.cell(row=carne_row, column=price_col).value
            price_usd_kg  = (float(price_per_ton) / 1000.0) if price_per_ton else None
        elif usd_col is not None and vol_tons:
            # Fallback: compute from US$ Mil / Toneladas
            usd_mil_val   = ws.cell(row=carne_row, column=usd_col).value
            price_usd_kg  = (float(usd_mil_val) / float(vol_tons)) if usd_mil_val else None
        else:
            price_usd_kg = None

        print(f"  [BULLETIN] Cols → ton={ton_col}, price={price_col}, usd={usd_col}  "
              f"| beef row={carne_row}")
        price_cell_val = ws.cell(row=carne_row, column=price_col).value if price_col else None
        print(f"  [BULLETIN] Extracted: vol={vol_tons!r}  price_per_ton={price_cell_val!r}")

        return (float(vol_tons) if vol_tons is not None else None), price_usd_kg

    vol_mtd    = None
    price_usd  = None
    used_sheet = None
    for ws in sheets_to_try:
        v, p = _parse_sheet(ws)
        if v is not None:
            vol_mtd    = v
            price_usd  = p
            used_sheet = ws.title
            print(f"  [BULLETIN] Sheet '{used_sheet}': vol_MTD={vol_mtd:,.0f} t"
                  + (f", price={price_usd:.4f} USD/kg" if price_usd else " (no price)"))
            break

    if vol_mtd is None:
        print("  [BULLETIN] Could not parse Excel. Sheet headers dump:")
        for ws in sheets_to_try[:2]:
            print(f"  Sheet: '{ws.title}'")
            for i, row in enumerate(ws.iter_rows(max_row=10)):
                vals = [(c.column, str(c.value or "")[:25]) for c in row if c.value]
                if vals:
                    print(f"    row {i+1}: {vals}")
        return 0

    # ── Sanity check on price ─────────────────────────────────────────────────
    PRICE_MIN, PRICE_MAX = 3.0, 20.0
    if price_usd is not None and not (PRICE_MIN <= price_usd <= PRICE_MAX):
        print(f"  [BULLETIN] Price {price_usd:.4f} out of range [{PRICE_MIN},{PRICE_MAX}] — discarding.")
        price_usd = None

    # ── Step 5: find the right _weekly_raw row to update ─────────────────────
    # Always update the latest row for the current month
    target_mo, target_yr = mo, yr
    yr_s = str(target_yr)
    mo_s = f"{target_mo:02d}"

    existing = conn.execute(
        "SELECT start_date, end_date, price_usd_kg FROM _weekly_raw"
        " WHERE start_date LIKE ? ORDER BY start_date DESC LIMIT 1",
        (f"{yr_s}-{mo_s}-%",)
    ).fetchone()

    if existing is None:
        # No row for this month yet — create one for the current week
        wk_start = today - _td(days=today.weekday())   # Monday
        s_date   = str(wk_start)
        e_date   = str(today)
        existing_price = None
    else:
        s_date, e_date, existing_price = existing

    # Validate existing_price too — don't preserve a previously bad value
    if existing_price is not None and not (PRICE_MIN <= existing_price <= PRICE_MAX):
        print(f"  [BULLETIN] Existing price {existing_price:.4f} also out of range — clearing.")
        existing_price = None

    # Keep the existing (good) price if our computed price failed the sanity check
    final_price = price_usd if price_usd is not None else existing_price

    conn.execute(
        "INSERT OR REPLACE INTO _weekly_raw(start_date,end_date,price_usd_kg,vol_tons)"
        " VALUES(?,?,?,?)",
        (s_date, e_date,
         round(final_price, 6) if final_price is not None else None,
         vol_mtd)
    )
    conn.commit()
    print(f"  [BULLETIN] _weekly_raw updated: {s_date} → {e_date} | "
          f"vol_MTD={vol_mtd:,.0f} t"
          + (f" | price={final_price:.4f} USD/kg" if final_price else ""))
    return 1


# ══════════════════════════════════════════════════════════════════════════════
# FILL MISSING SECEX MONTHS FROM WEEKLY AVERAGES
# ══════════════════════════════════════════════════════════════════════════════
def fill_secex_from_weekly(conn):
    """
    For months after the last official SECEX entry, estimate price_usd_kg as the
    simple average of weekly prices that fall within that month.

    Uses INSERT OR IGNORE so that when real MDIC data arrives via fetch_secex()
    (which uses INSERT OR REPLACE), the official values automatically overwrite
    these estimates.

    Returns the number of newly inserted estimated rows.
    """
    last = conn.execute(
        "SELECT year, month FROM _secex_raw ORDER BY year DESC, month DESC LIMIT 1"
    ).fetchone()
    if not last:
        return 0

    ly, lm = last
    last_date = f"{ly}-{lm:02d}-{monthrange(ly, lm)[1]:02d}"

    weekly = conn.execute(
        """
        SELECT CAST(strftime('%Y', start_date) AS INTEGER),
               CAST(strftime('%m', start_date) AS INTEGER),
               AVG(price_usd_kg)
        FROM   _weekly_raw
        WHERE  start_date > ? AND price_usd_kg IS NOT NULL
        GROUP  BY 1, 2
        ORDER  BY 1, 2
        """,
        (last_date,),
    ).fetchall()

    filled = 0
    for yr, mo, avg_p in weekly:
        if avg_p is None:
            continue
        conn.execute(
            "INSERT OR IGNORE INTO _secex_raw(year,month,rev_000usd,vol_tons,price_usd_kg)"
            " VALUES(?,?,NULL,NULL,?)",
            (yr, mo, round(avg_p, 6)),
        )
        if conn.execute("SELECT changes()").fetchone()[0]:
            print(f"  [SECEX-est] {yr}-{mo:02d}: {avg_p:.4f} USD/kg  ← weekly avg (official MDIC pending)")
            filled += 1

    conn.commit()
    return filled


# ══════════════════════════════════════════════════════════════════════════════
# COMPUTE & MATERIALISE SPREAD TABLES
# ══════════════════════════════════════════════════════════════════════════════
def _avg(conn, table, col, s, e):
    r = conn.execute(
        f"SELECT AVG({col}) FROM {table} WHERE dt >= ? AND dt <= ?", (s, e)
    ).fetchone()
    return r[0] if r and r[0] is not None else None


def materialise(conn):
    """Compute monthly and weekly spread tables from raw data."""
    now_iso = datetime.now().isoformat(timespec="seconds")

    # ── Monthly ───────────────────────────────────────────────────────────────
    raw_m = conn.execute(
        "SELECT year, month, price_usd_kg FROM _secex_raw ORDER BY year, month"
    ).fetchall()
    monthly_rows = []
    for yr, mo, p_usd in raw_m:
        s  = f"{yr}-{mo:02d}-01"
        ld = monthrange(yr, mo)[1]
        e  = f"{yr}-{mo:02d}-{ld:02d}"
        fx = _avg(conn, "_fx_raw", "fx", s, e)
        ca = _avg(conn, "_cepea_raw", "r_kg", s, e)
        brl = (p_usd * fx) if p_usd and fx else None
        sp  = (brl / ca)  if brl and ca  else None
        monthly_rows.append((
            f"{yr}-{mo:02d}", yr, mo,
            round(p_usd, 6) if p_usd else None,
            round(fx,   6) if fx else None,
            round(brl,  6) if brl else None,
            round(ca,   6) if ca else None,
            round(sp,   6) if sp else None,
            now_iso,
        ))
    conn.executemany(
        "INSERT OR REPLACE INTO monthly"
        "(period,year,month,secex_usd_kg,fx,secex_brl_kg,cepea_r_kg,spread,updated_at)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        monthly_rows
    )

    # ── Weekly ────────────────────────────────────────────────────────────────
    raw_w = conn.execute(
        "SELECT start_date, end_date, price_usd_kg, vol_tons"
        " FROM _weekly_raw ORDER BY start_date"
    ).fetchall()

    # De-accumulate MTD vol_tons → incremental weekly vol_tons
    # SECEX bulletin reports cumulative tonnage since start of month (MTD).
    # For each week: vol_week = vol_MTD − vol_MTD_of_previous_week_in_same_month.
    # First week of each month: vol_week = vol_MTD (nothing to subtract).
    prev_mtd: dict[tuple, float | None] = {}  # key = (year, month)

    weekly_rows = []
    for s, e, p_usd, vol_mtd in raw_w:
        fx  = _avg(conn, "_fx_raw",   "fx",  s, e)
        ca  = _avg(conn, "_cepea_raw", "r_kg", s, e)
        brl = (p_usd * fx) if p_usd and fx else None
        sp  = (brl / ca)  if brl and ca   else None

        # De-accumulate volume
        yr_mo = (int(s[:4]), int(s[5:7]))
        if vol_mtd is not None:
            prev = prev_mtd.get(yr_mo)
            vol_week = vol_mtd - prev if prev is not None else vol_mtd
            prev_mtd[yr_mo] = vol_mtd
        else:
            vol_week = None
            # Leave prev_mtd unchanged so next week still de-accumulates correctly

        weekly_rows.append((
            s, e,
            round(p_usd,    6) if p_usd    else None,
            round(fx,       6) if fx       else None,
            round(brl,      6) if brl      else None,
            round(ca,       6) if ca       else None,
            round(sp,       6) if sp       else None,
            round(vol_week, 3) if vol_week is not None else None,
            now_iso,
        ))

    conn.executemany(
        "INSERT OR REPLACE INTO weekly"
        "(start_date,end_date,secex_usd_kg,fx,secex_brl_kg,cepea_r_kg,spread,vol_tons,updated_at)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        weekly_rows
    )
    conn.commit()

    nw_vol = sum(1 for r in weekly_rows if r[7] is not None)
    nm = conn.execute("SELECT COUNT(*) FROM monthly WHERE spread IS NOT NULL").fetchone()[0]
    nw = conn.execute("SELECT COUNT(*) FROM weekly  WHERE spread IS NOT NULL").fetchone()[0]
    print(f"  [DB] monthly: {len(monthly_rows)} rows ({nm} with spread)")
    print(f"  [DB] weekly:  {len(weekly_rows)} rows ({nw} with spread, {nw_vol} with volume)")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    args = sys.argv[1:]
    do_init    = "--init" in args
    cepea_idx  = next((i for i, a in enumerate(args) if a == "--cepea"), None)
    cepea_path = args[cepea_idx + 1] if cepea_idx is not None and cepea_idx + 1 < len(args) else None

    print("=" * 60)
    print(f"  Brazil Beef Spread Extractor — {date.today().isoformat()}")
    print(f"  Mode: {'INIT (full seed)' if do_init else 'INCREMENTAL UPDATE'}")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    # Weekly raw always seeded (INSERT OR REPLACE → idempotent)
    print("\n[1] Seeding weekly raw data …")
    seed_weekly_raw(conn)

    if do_init:
        # Full SECEX history
        print("\n[2] Fetching SECEX monthly (all years) …")
        fetch_secex(conn)

        # CEPEA from local XLS
        print("\n[3] Loading CEPEA from local XLS …")
        if cepea_path:
            load_cepea_xls(conn, cepea_path)
        else:
            candidates = (
                list(Path(DB_PATH.parent).glob("CEPEA*.xls*")) +
                list(Path(DB_PATH.parent).glob("cepea*.xls*")) +
                list(Path(DB_PATH.parent.parent).glob("CEPEA*.xls*"))
            )
            if candidates:
                load_cepea_xls(conn, candidates[0])
            else:
                print("  [CEPEA] No XLS file found.")
                print("  Run with:  python extractor_bz.py --init --cepea /path/to/CEPEA.xls")
                print("  Download:  https://www.cepea.esalq.usp.br/br/indicador/boi-gordo.aspx")

        print("\n[4] Fetching BCB PTAX FX …")
        fetch_fx(conn)

    else:
        # Incremental: current + prior year SECEX only
        print("\n[2] Fetching SECEX monthly (recent years) …")
        yr = datetime.now().year
        fetch_secex(conn, years=[yr - 1, yr])

        print("\n[3] Fetching BCB PTAX FX …")
        fetch_fx(conn)

    # Weekly bulletin: fetch latest price + MTD volume (both init and incremental)
    print("\n[→] Fetching SECEX weekly bulletin (price + MTD volume) …")
    fetch_weekly_bulletin(conn)

    print("\n[→] Computing spread tables …")
    materialise(conn)

    conn.close()
    print("\n" + "=" * 60)
    print("  Done. beef_bz.db updated.")
    print("=" * 60)


if __name__ == "__main__":
    main()
