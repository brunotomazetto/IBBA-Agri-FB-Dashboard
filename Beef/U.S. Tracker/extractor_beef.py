#!/usr/bin/env python3
"""
extractor_beef.py
─────────────────
Creates and maintains beef.db (SQLite) — the data backend for the
U.S. Beef Packer Margin Tracker dashboard.

USAGE
-----
  # One-time historical load from the V4 Excel tracker:
  python extractor_beef.py --history "US_Beef_Packer_Margin_Tracker_v4.xlsx"

  # Weekly update (run every Monday via cron / GitHub Actions):
  python extractor_beef.py

  # Weekly update AND rebuild all quarterly averages from scratch:
  python extractor_beef.py --full

DEPENDENCIES
------------
  pip install requests pandas openpyxl pdfplumber

DATA SOURCES
------------
  Historical  → US_Beef_Packer_Margin_Tracker_v4.xlsx (local file)
  CT150       → USDA AMS Datamart API, report 2461
  Cutout      → USDA AMS Datamart API, report 2527
  KS prices   → USDA AMS PDF, report 2484
  NE prices   → USDA AMS PDF, report 2667
  Drop Credit → USDA AMS PDF, report 2829
  Slaughter   → USDA NASS Quick Stats API
  Henry Hub   → EIA Open Data API, series NG.RNGWHHD.D

OUTPUT
------
  beef.db  (SQLite, same folder as this script)
  → beef_weekly    : one row per week
  → beef_quarterly : one row per quarter (market averages + company financials)
"""

import argparse
import io
import math
import re
import sqlite3
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import requests
import pandas as pd

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False
    print("⚠  pdfplumber not installed — KS/NE and Drop Credit will be skipped.\n"
          "   Run: pip install pdfplumber")

# ─── CONFIG ──────────────────────────────────────────────────────────────────
DB_PATH      = Path(__file__).parent / "beef.db"
import os
EIA_API_KEY  = os.environ.get('EIA_API_KEY', 'YOUR_EIA_API_KEY')
LOOKBACK_WEEKS = 6
# ─────────────────────────────────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE IF NOT EXISTS beef_weekly (
    week_ending  TEXT PRIMARY KEY,
    slaughter    REAL,
    ct150_steer  REAL, ct150_heifer REAL, ct150_mixed REAL, ct150_all REAL,
    ks_steer     REAL, ks_heifer    REAL, ks_avg      REAL,
    ne_steer     REAL, ne_heifer    REAL, ne_avg      REAL,
    choice       REAL, select_      REAL,
    drop_credit  REAL,
    henry_hub    REAL,
    updated_at   TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS beef_quarterly (
    quarter          TEXT PRIMARY KEY,
    quarter_start    TEXT,
    slaughter        REAL,
    ct150_steer      REAL, ct150_heifer REAL, ct150_mixed REAL, ct150_all REAL,
    ks_steer         REAL, ks_heifer    REAL, ks_avg      REAL,
    ne_steer         REAL, ne_heifer    REAL, ne_avg      REAL,
    choice           REAL, select_      REAL,
    drop_credit      REAL, henry_hub    REAL,
    -- Marfrig / National Beef
    mbrf_revenue     REAL, mbrf_gp      REAL, mbrf_gm         REAL,
    mbrf_ebitda      REAL, mbrf_ebitda_mgn REAL,
    -- JBS North America
    jbs_revenue      REAL, jbs_gp       REAL, jbs_gm          REAL,
    jbs_ebit         REAL, jbs_ebit_mgn REAL,
    jbs_ebitda       REAL, jbs_ebitda_mgn REAL,
    -- Tyson Beef
    tyson_sales      REAL, tyson_adj_op_inc REAL, tyson_adj_op_mgn REAL,
    updated_at       TEXT DEFAULT (datetime('now'))
);
"""

MARKET_COLS = [
    "slaughter",
    "ct150_steer","ct150_heifer","ct150_mixed","ct150_all",
    "ks_steer","ks_heifer","ks_avg",
    "ne_steer","ne_heifer","ne_avg",
    "choice","select_","drop_credit","henry_hub",
]
FIN_COLS = [
    "mbrf_revenue","mbrf_gp","mbrf_gm","mbrf_ebitda","mbrf_ebitda_mgn",
    "jbs_revenue","jbs_gp","jbs_gm","jbs_ebit","jbs_ebit_mgn",
    "jbs_ebitda","jbs_ebitda_mgn",
    "tyson_sales","tyson_adj_op_inc","tyson_adj_op_mgn",
]

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def fv(v):
    """Float or None."""
    if v is None or v == "": return None
    try:
        f = float(str(v).replace(",", ""))
        return None if math.isnan(f) else round(f, 6)
    except (TypeError, ValueError):
        return None

def dv(v):
    """Date → ISO string or None."""
    if v is None: return None
    if isinstance(v, date): return v.isoformat()
    try:
        return pd.to_datetime(v).date().isoformat()
    except Exception:
        return None

def week_end_sat(d: date) -> date:
    return d + timedelta(days=(5 - d.weekday()) % 7)

def quarter_label(d: date) -> str:
    q = (d.month - 1) // 3 + 1
    return f"{q}Q{str(d.year)[2:]}"

def quarter_start(q: str) -> str:
    """'1Q18' → '2018-01-01'"""
    m = re.match(r"([1-4])Q(\d{2})", q)
    if not m: return None
    qn, yr = int(m.group(1)), 2000 + int(m.group(2))
    return date(yr, (qn - 1) * 3 + 1, 1).isoformat()

def since_date() -> str:
    return (date.today() - timedelta(weeks=LOOKBACK_WEEKS)).isoformat()

def _parse_ams_date(text: str) -> date:
    """Extract week-ending date from AMS PDF text; falls back to this Saturday."""
    patterns = [
        r"[Ww]eek\s+[Ee]nd(?:ing)?\s*[:\-]?\s*(\w+ \d{1,2},?\s+\d{4})",
        r"[Ff]or (?:the )?[Ww]eek (?:of\s+)?(\w+ \d{1,2},?\s+\d{4})",
        r"[Ww]eek\s+[Ee]nd(?:ing)?\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})",
        r"(\w{3,9} \d{1,2},?\s+\d{4})",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            try:
                return week_end_sat(pd.to_datetime(m.group(1)).date())
            except Exception:
                continue
    return week_end_sat(date.today())

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════

def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    conn.commit()
    return conn

def upsert_weekly(conn: sqlite3.Connection, rows: list[dict]):
    if not rows: return
    cols = [c for c in ["week_ending","slaughter",
                        "ct150_steer","ct150_heifer","ct150_mixed","ct150_all",
                        "ks_steer","ks_heifer","ks_avg",
                        "ne_steer","ne_heifer","ne_avg",
                        "choice","select_","drop_credit","henry_hub"]
            if any(c in r for r in rows)]
    all_cols = list({c for r in rows for c in r})
    sql = (f"INSERT INTO beef_weekly ({','.join(all_cols)}) "
           f"VALUES ({','.join('?' for _ in all_cols)}) "
           f"ON CONFLICT(week_ending) DO UPDATE SET "
           + ", ".join(f"{c}=excluded.{c}" for c in all_cols if c != "week_ending"))
    conn.executemany(sql, [[r.get(c) for c in all_cols] for r in rows])
    conn.commit()

def upsert_quarterly(conn: sqlite3.Connection, rows: list[dict]):
    if not rows: return
    all_cols = list({c for r in rows for c in r})
    sql = (f"INSERT INTO beef_quarterly ({','.join(all_cols)}) "
           f"VALUES ({','.join('?' for _ in all_cols)}) "
           f"ON CONFLICT(quarter) DO UPDATE SET "
           + ", ".join(f"{c}=excluded.{c}" for c in all_cols if c != "quarter"))
    conn.executemany(sql, [[r.get(c) for c in all_cols] for r in rows])
    conn.commit()

# ══════════════════════════════════════════════════════════════════════════════
# HISTORICAL LOAD FROM V4 EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def load_history(tracker_path: str, conn: sqlite3.Connection):
    """
    Reads both Summary sheets from the V4 tracker and populates beef.db.
    Column mapping (1-indexed):
      Weekly/Quarterly cols 1-16: date/quarter, slaughter, CT150 x4,
                                   KS steer/heifer/avg, NE steer/heifer/avg,
                                   choice, select, drop_credit, henry_hub
      Quarterly cols 18-32: MBRF (5 cols), JBS (7 cols), Tyson (3 cols)
    """
    import openpyxl
    print(f"\nLoading workbook: {tracker_path}")
    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    # ── Weekly ──────────────────────────────────────────────────────────────
    ws = wb["Summary \u2013 Weekly"]
    weekly_rows = []
    for r in range(5, ws.max_row + 1):
        dt = dv(ws.cell(r, 1).value)
        if not dt: continue
        weekly_rows.append({
            "week_ending":  dt,
            "slaughter":    fv(ws.cell(r, 2).value),
            "ct150_steer":  fv(ws.cell(r, 3).value),
            "ct150_heifer": fv(ws.cell(r, 4).value),
            "ct150_mixed":  fv(ws.cell(r, 5).value),
            "ct150_all":    fv(ws.cell(r, 6).value),
            "ks_steer":     fv(ws.cell(r, 7).value),
            "ks_heifer":    fv(ws.cell(r, 8).value),
            "ks_avg":       fv(ws.cell(r, 9).value),
            "ne_steer":     fv(ws.cell(r,10).value),
            "ne_heifer":    fv(ws.cell(r,11).value),
            "ne_avg":       fv(ws.cell(r,12).value),
            "choice":       fv(ws.cell(r,13).value),
            "select_":      fv(ws.cell(r,14).value),
            "drop_credit":  fv(ws.cell(r,15).value),
            "henry_hub":    fv(ws.cell(r,16).value),
        })
    print(f"  Weekly rows: {len(weekly_rows)}")
    upsert_weekly(conn, weekly_rows)

    # ── Quarterly ────────────────────────────────────────────────────────────
    ws = wb["Summary \u2013 Quarterly"]
    qtr_rows = []
    for r in range(5, ws.max_row + 1):
        q = ws.cell(r, 1).value
        if not q or not re.match(r"[1-4]Q\d{2}", str(q)): continue
        q = str(q)
        qtr_rows.append({
            "quarter":          q,
            "quarter_start":    quarter_start(q),
            "slaughter":        fv(ws.cell(r, 2).value),
            "ct150_steer":      fv(ws.cell(r, 3).value),
            "ct150_heifer":     fv(ws.cell(r, 4).value),
            "ct150_mixed":      fv(ws.cell(r, 5).value),
            "ct150_all":        fv(ws.cell(r, 6).value),
            "ks_steer":         fv(ws.cell(r, 7).value),
            "ks_heifer":        fv(ws.cell(r, 8).value),
            "ks_avg":           fv(ws.cell(r, 9).value),
            "ne_steer":         fv(ws.cell(r,10).value),
            "ne_heifer":        fv(ws.cell(r,11).value),
            "ne_avg":           fv(ws.cell(r,12).value),
            "choice":           fv(ws.cell(r,13).value),
            "select_":          fv(ws.cell(r,14).value),
            "drop_credit":      fv(ws.cell(r,15).value),
            "henry_hub":        fv(ws.cell(r,16).value),
            # col 17 = spacer
            "mbrf_revenue":     fv(ws.cell(r,18).value),
            "mbrf_gp":          fv(ws.cell(r,19).value),
            "mbrf_gm":          fv(ws.cell(r,20).value),
            "mbrf_ebitda":      fv(ws.cell(r,21).value),
            "mbrf_ebitda_mgn":  fv(ws.cell(r,22).value),
            "jbs_revenue":      fv(ws.cell(r,23).value),
            "jbs_gp":           fv(ws.cell(r,24).value),
            "jbs_gm":           fv(ws.cell(r,25).value),
            "jbs_ebit":         fv(ws.cell(r,26).value),
            "jbs_ebit_mgn":     fv(ws.cell(r,27).value),
            "jbs_ebitda":       fv(ws.cell(r,28).value),
            "jbs_ebitda_mgn":   fv(ws.cell(r,29).value),
            "tyson_sales":      fv(ws.cell(r,30).value),
            "tyson_adj_op_inc": fv(ws.cell(r,31).value),
            "tyson_adj_op_mgn": fv(ws.cell(r,32).value),
        })
    print(f"  Quarterly rows: {len(qtr_rows)}")
    upsert_quarterly(conn, qtr_rows)
    print("✓ Historical load complete.")

# ══════════════════════════════════════════════════════════════════════════════
# FETCH FUNCTIONS — USDA AMS + EIA
# ══════════════════════════════════════════════════════════════════════════════

def fetch_ct150() -> pd.DataFrame:
    """CT150 5-area weighted average — USDA AMS Datamart report 2461."""
    url = "https://marsapi.ams.usda.gov/services/v1.2/reports/2461"
    r = requests.get(url, params={"q": f"report_date>={since_date()}", "allSections": "true"}, timeout=30)
    r.raise_for_status()
    CLASS_MAP = {"Steers": "ct150_steer", "Heifers": "ct150_heifer",
                 "Mixed":  "ct150_mixed", "All Beef": "ct150_all"}
    rows: dict[date, dict] = defaultdict(dict)
    for rec in r.json().get("results", []):
        raw = pd.to_datetime(rec.get("report_date")).date()
        sat = week_end_sat(raw - timedelta(days=2))  # Monday report → Saturday
        col = CLASS_MAP.get(rec.get("class_description", ""))
        val = fv(rec.get("wtd_avg"))
        if col and val is not None:
            rows[sat][col] = val
    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "week_ending"
    print(f"  CT150: {len(df)} weeks")
    return df.reset_index()

def fetch_cutout() -> pd.DataFrame:
    """Choice & Select cutout — USDA AMS Datamart report 2527."""
    url = "https://marsapi.ams.usda.gov/services/v1.2/reports/2527"
    r = requests.get(url, params={"q": f"report_date>={since_date()}", "allSections": "true"}, timeout=30)
    r.raise_for_status()
    GRADE_MAP = {"Choice 600-900": "choice", "Select 600-900": "select_"}
    rows: dict[date, dict] = defaultdict(dict)
    for rec in r.json().get("results", []):
        raw = pd.to_datetime(rec.get("report_date")).date()
        sat = week_end_sat(raw + timedelta(days=1))  # Friday report → Saturday
        grade = (rec.get("grade", "") + " " + rec.get("weight_range", "")).strip()
        col = GRADE_MAP.get(grade)
        val = fv(rec.get("avg_price"))
        if col and val is not None:
            rows[sat][col] = val
    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "week_ending"
    print(f"  Cutout: {len(df)} weeks")
    return df.reset_index()

def fetch_slaughter() -> pd.DataFrame:
    """Weekly cattle slaughter — USDA NASS Quick Stats API."""
    NASS_KEY = "YOUR_NASS_API_KEY"   # https://quickstats.nass.usda.gov/api
    url = "https://quickstats.nass.usda.gov/api/api_GET/"
    params = {
        "key": NASS_KEY, "commodity_desc": "CATTLE",
        "statisticcat_desc": "SLAUGHTER", "unit_desc": "HEAD",
        "freq_desc": "WEEKLY", "begin_DT": since_date(), "format": "JSON",
    }
    try:
        r = requests.get(url, params=params, timeout=30)
        r.raise_for_status()
        data = r.json().get("data", [])
    except Exception as e:
        print(f"  Slaughter fetch failed: {e}"); return pd.DataFrame()
    rows = {}
    for rec in data:
        raw = rec.get("week_ending") or rec.get("end_code", "")
        if not raw: continue
        try: d = pd.to_datetime(raw).date()
        except: continue
        v = fv(rec.get("Value", "").replace(",", ""))
        if v: rows[week_end_sat(d)] = v
    df = pd.DataFrame(list(rows.items()), columns=["week_ending", "slaughter"])
    print(f"  Slaughter: {len(df)} weeks")
    return df

def fetch_ks_ne() -> pd.DataFrame:
    """
    KS prices: USDA AMS report 2484  (mymarketnews.ams.usda.gov/viewReport/2484)
    NE prices: USDA AMS report 2667  (mymarketnews.ams.usda.gov/viewReport/2667)

    Strategy: try MARS JSON API first; fall back to PDF parsing.
    Parses "WEEKLY ACCUMULATED" section → Live Steer / Heifer Avg Price.
    """
    REPORTS = {
        "ks": {"mars_id": "2484", "pdf_slug": "ams_2484"},
        "ne": {"mars_id": "2667", "pdf_slug": "ams_2667"},
    }
    rows: dict[date, dict] = {}

    for state, cfg in REPORTS.items():
        steer_col  = f"{state}_steer"
        heifer_col = f"{state}_heifer"
        fetched    = False

        # ── MARS API ──────────────────────────────────────────────────────────
        try:
            url = f"https://marsapi.ams.usda.gov/services/v1.2/reports/{cfg['mars_id']}"
            r = requests.get(url, params={"q": f"report_date>={since_date()}", "allSections": "true"}, timeout=30)
            r.raise_for_status()
            for rec in r.json().get("results", []):
                section = str(rec.get("report_section", "")).lower()
                class_  = str(rec.get("class_description", "")).lower()
                if "accum" not in section and "weekly" not in section: continue
                if "live" not in str(rec.get("type_description", "")).lower(): continue
                price = fv(rec.get("wtd_avg") or rec.get("avg_price"))
                if price is None: continue
                sat = week_end_sat(pd.to_datetime(rec.get("report_date")).date())
                rows.setdefault(sat, {})
                if "steer"  in class_: rows[sat][steer_col]  = price; fetched = True
                if "heifer" in class_: rows[sat][heifer_col] = price; fetched = True
            if fetched:
                print(f"  {state.upper()}: fetched via MARS API"); continue
        except Exception as e:
            print(f"  {state.upper()} MARS API error ({e}); trying PDF…")

        # ── PDF fallback ───────────────────────────────────────────────────────
        if not HAS_PDFPLUMBER:
            print(f"  {state.upper()}: pdfplumber not available — skipping"); continue
        try:
            pdf_url = f"https://www.ams.usda.gov/mnreports/{cfg['pdf_slug']}.pdf"
            resp = requests.get(pdf_url, timeout=30)
            resp.raise_for_status()
            with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            sat = _parse_ams_date(text)
            acc_idx = text.upper().find("WEEKLY ACCUMULATED")
            if acc_idx == -1:
                print(f"  {state.upper()}: 'WEEKLY ACCUMULATED' not found"); continue
            section = text[acc_idx : acc_idx + 700]
            rows.setdefault(sat, {})
            for animal, col in [("Steer", steer_col), ("Heifer", heifer_col)]:
                m = re.search(r"Live\s+" + animal + r"\s+[\d,]+\s+[\d,.]+\s+\$?([\d.]+)",
                              section, re.IGNORECASE)
                if m: rows[sat][col] = fv(m.group(1))
            print(f"  {state.upper()}: steer=${rows[sat].get(steer_col)}, "
                  f"heifer=${rows[sat].get(heifer_col)}  (PDF, week {sat})")
        except Exception as e:
            print(f"  {state.upper()} PDF error: {e}")

    if not rows: return pd.DataFrame()
    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "week_ending"
    for st in ["ks", "ne"]:
        sc, hc, ac = f"{st}_steer", f"{st}_heifer", f"{st}_avg"
        if sc in df.columns and hc in df.columns:
            df[ac] = df[[sc, hc]].mean(axis=1, skipna=False)
    return df.reset_index()

def fetch_drop_credit() -> pd.DataFrame:
    """
    By-product value — USDA AMS report 2829
    PDF: https://www.ams.usda.gov/mnreports/ams_2829.pdf
    Extracts "Dressed Equivalent Basis" from the STEER table ($/cwt dressed).
    """
    if not HAS_PDFPLUMBER:
        print("  Drop Credit: pdfplumber not available — skipping")
        return pd.DataFrame()
    try:
        r = requests.get("https://www.ams.usda.gov/mnreports/ams_2829.pdf", timeout=30)
        r.raise_for_status()
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        print(f"  Drop Credit fetch error: {e}"); return pd.DataFrame()

    sat = _parse_ams_date(text)
    drop_val = None
    m = re.search(r"[Dd]ressed\s+[Ee]quivalent\s+[Bb]asis\s*\([^)]+\)\s*[:\s]+([\d.]+)", text)
    if m:
        drop_val = fv(m.group(1))
        print(f"  Drop Credit: {drop_val} $/cwt dressed  (week {sat})")
    else:
        m = re.search(r"[Tt]otals?\s*:?\s*[\d.]+\s+[\d.]+\s+([\d.]+)\s*$", text, re.MULTILINE)
        if m:
            drop_val = fv(m.group(1))
            print(f"  Drop Credit: {drop_val} $/cwt live (Totals fallback, week {sat})")
    if drop_val is None:
        print("  Drop Credit: could not parse value"); return pd.DataFrame()
    return pd.DataFrame([{"week_ending": sat, "drop_credit": drop_val}])

def fetch_henry_hub() -> pd.DataFrame:
    """Henry Hub daily spot → weekly average — EIA Open Data API."""
    try:
        r = requests.get(
            "https://api.eia.gov/v2/natural-gas/pri/sum/data/",
            params={
                "api_key": EIA_API_KEY, "frequency": "daily",
                "data[0]": "value", "facets[series][]": "NG.RNGWHHD.D",
                "start": since_date(), "sort[0][column]": "period",
                "sort[0][direction]": "asc", "length": 500,
            }, timeout=30
        )
        r.raise_for_status()
        records = r.json().get("response", {}).get("data", [])
    except Exception as e:
        print(f"  Henry Hub fetch failed: {e}"); return pd.DataFrame()
    daily = {}
    for rec in records:
        try: d = pd.to_datetime(rec["period"]).date()
        except: continue
        v = fv(rec.get("value"))
        if v: daily[d] = v
    weekly: dict[date, list] = defaultdict(list)
    for d, v in daily.items():
        weekly[week_end_sat(d)].append(v)
    rows = [{"week_ending": sat, "henry_hub": round(sum(vs)/len(vs), 4)}
            for sat, vs in weekly.items() if vs]
    df = pd.DataFrame(rows)
    print(f"  Henry Hub: {len(df)} weeks")
    return df

# ══════════════════════════════════════════════════════════════════════════════
# WEEKLY UPDATE
# ══════════════════════════════════════════════════════════════════════════════

def update_weekly(conn: sqlite3.Connection):
    print(f"\n=== Fetching weekly data (last {LOOKBACK_WEEKS} weeks) ===")
    dfs = []
    for name, fn in [("CT150", fetch_ct150), ("Cutout", fetch_cutout),
                     ("Slaughter", fetch_slaughter), ("KS-NE", fetch_ks_ne),
                     ("Drop", fetch_drop_credit), ("HH", fetch_henry_hub)]:
        try:
            df = fn()
            if not df.empty:
                df["week_ending"] = pd.to_datetime(df["week_ending"]).dt.date.apply(
                    lambda d: d.isoformat() if d else None)
                dfs.append(df.set_index("week_ending"))
        except Exception as e:
            print(f"  {name} error: {e}")

    if not dfs:
        print("No weekly data to write."); return

    merged = dfs[0]
    for df in dfs[1:]:
        merged = merged.join(df, how="outer")

    # Compute KS / NE averages if individual steer/heifer present
    for st in ["ks", "ne"]:
        sc, hc, ac = f"{st}_steer", f"{st}_heifer", f"{st}_avg"
        if sc in merged.columns and hc in merged.columns and ac not in merged.columns:
            merged[ac] = merged[[sc, hc]].mean(axis=1, skipna=False)

    merged = merged.reset_index()
    merged = merged.where(pd.notnull(merged), None)
    rows = merged.to_dict("records")
    upsert_weekly(conn, rows)
    print(f"  ✓ beef_weekly: {len(rows)} rows upserted")

# ══════════════════════════════════════════════════════════════════════════════
# QUARTERLY RECOMPUTE
# ══════════════════════════════════════════════════════════════════════════════

def recompute_quarterly(conn: sqlite3.Connection, full: bool = False):
    print("\n=== Recomputing quarterly averages ===")
    if full:
        rows = conn.execute("SELECT * FROM beef_weekly").fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM beef_weekly WHERE week_ending >= ?", (since_date(),)
        ).fetchall()

    if not rows:
        print("  No weekly rows found."); return

    df = pd.DataFrame([dict(r) for r in rows])
    df["week_ending"] = pd.to_datetime(df["week_ending"]).dt.date
    df["quarter"]     = df["week_ending"].apply(quarter_label)

    agg = df.groupby("quarter")[MARKET_COLS].mean().reset_index()
    agg = agg.where(pd.notnull(agg), None)
    agg["quarter_start"] = agg["quarter"].apply(quarter_start)

    # Preserve existing company financials
    quarters = agg["quarter"].tolist()
    fin_rows = conn.execute(
        f"SELECT * FROM beef_quarterly WHERE quarter IN ({','.join('?'*len(quarters))})",
        quarters
    ).fetchall()
    fin_map = {r["quarter"]: dict(r) for r in fin_rows}

    out_rows = []
    for _, row in agg.iterrows():
        q = row["quarter"]
        rec = {"quarter": q, "quarter_start": row["quarter_start"]}
        for col in MARKET_COLS:
            v = row.get(col)
            rec[col] = round(float(v), 6) if v is not None else None
        existing = fin_map.get(q, {})
        for col in FIN_COLS:
            rec[col] = existing.get(col)
        out_rows.append(rec)

    upsert_quarterly(conn, out_rows)
    print(f"  ✓ beef_quarterly: {len(out_rows)} rows recomputed")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="U.S. Beef Tracker — SQLite extractor")
    parser.add_argument("--history", metavar="XLSX_PATH",
                        help="Path to V4 tracker (.xlsx) — runs one-time historical load")
    parser.add_argument("--full", action="store_true",
                        help="Rebuild ALL quarterly rows from beef_weekly (use with weekly update)")
    args = parser.parse_args()

    conn = init_db()
    print(f"Database: {DB_PATH}")

    if args.history:
        load_history(args.history, conn)
    else:
        update_weekly(conn)
        recompute_quarterly(conn, full=args.full)

    conn.close()
    print("\n✓ Done.")
