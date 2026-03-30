"""
load_history.py
───────────────
One-time script: reads US_Beef_Packer_Margin_Tracker_v4.xlsx and uploads
ALL historical data to Supabase (beef_weekly + beef_quarterly tables).

Usage:
    pip install supabase openpyxl
    python load_history.py

Configure the three constants below before running.
"""

import openpyxl
from datetime import date, timedelta
from supabase import create_client, Client

# ─── CONFIG ──────────────────────────────────────────────────────────────────
SUPABASE_URL     = "https://vhxvlmyataclkpamvzpe.supabase.co"
SUPABASE_SERVICE  = "YOUR_SERVICE_ROLE_KEY"     # Settings → API → service_role (secret)
TRACKER_PATH     = "US_Beef_Packer_Margin_Tracker_v4.xlsx"   # adjust path if needed
BATCH_SIZE       = 200    # rows per Supabase upsert call
# ─────────────────────────────────────────────────────────────────────────────

sb: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE)

def fv(v):
    """Float or None."""
    if v is None or v == "": return None
    try:
        f = float(v)
        import math
        return None if math.isnan(f) else round(f, 6)
    except (TypeError, ValueError):
        return None

def dv(v):
    """Date value from cell (already a date object or None)."""
    if v is None: return None
    if isinstance(v, date): return v.isoformat()
    try:
        return v.date().isoformat()
    except Exception:
        return None

def batch_upsert(table: str, rows: list, conflict: str):
    """Upload rows in batches, using upsert on conflict."""
    for i in range(0, len(rows), BATCH_SIZE):
        chunk = rows[i:i+BATCH_SIZE]
        sb.table(table).upsert(chunk, on_conflict=conflict).execute()
        print(f"  [{table}] uploaded rows {i+1}–{min(i+BATCH_SIZE, len(rows))}")

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOAD WEEKLY DATA  (Summary – Weekly tab)
# ══════════════════════════════════════════════════════════════════════════════
# Column map (1-indexed, matches V4 tracker):
#   1=Week Ending, 2=Slaughter
#   3=CT150 Steer, 4=CT150 Heifer, 5=CT150 Mixed, 6=CT150 All Beef
#   7=KS Steer, 8=KS Heifer, 9=KS Avg
#   10=NE Steer, 11=NE Heifer, 12=NE Avg
#   13=Choice, 14=Select, 15=Drop Credit, 16=Henry Hub

def load_weekly(wb):
    ws = wb["Summary \u2013 Weekly"]
    rows = []
    for r in range(5, ws.max_row + 1):
        dt = dv(ws.cell(r, 1).value)
        if not dt:
            continue
        rows.append({
            "week_ending":  dt,
            "slaughter":    fv(ws.cell(r, 2).value),
            "ct150_steer":  fv(ws.cell(r, 3).value),
            "ct150_heifer": fv(ws.cell(r, 4).value),
            "ct150_mixed":  fv(ws.cell(r, 5).value),
            "ct150_all":    fv(ws.cell(r, 6).value),
            "ks_steer":     fv(ws.cell(r, 7).value),
            "ks_heifer":    fv(ws.cell(r, 8).value),
            "ks_avg":       fv(ws.cell(r, 9).value),
            "ne_steer":     fv(ws.cell(r,10).value),
            "ne_heifer":    fv(ws.cell(r,11).value),
            "ne_avg":       fv(ws.cell(r,12).value),
            "choice":       fv(ws.cell(r,13).value),
            "select_":      fv(ws.cell(r,14).value),
            "drop_credit":  fv(ws.cell(r,15).value),
            "henry_hub":    fv(ws.cell(r,16).value),
        })
    print(f"\nWeekly rows read: {len(rows)}")
    batch_upsert("beef_weekly", rows, "week_ending")

# ══════════════════════════════════════════════════════════════════════════════
# 2. LOAD QUARTERLY DATA  (Summary – Quarterly tab)
# ══════════════════════════════════════════════════════════════════════════════
# Column map (1-indexed, V4):
#   1=Quarter, 2=Slaughter
#   3=CT150 Steer, 4=CT150 Heifer, 5=CT150 Mixed, 6=CT150 All Beef
#   7=KS Steer, 8=KS Heifer, 9=KS Avg
#   10=NE Steer, 11=NE Heifer, 12=NE Avg
#   13=Choice, 14=Select, 15=Drop Credit, 16=Henry Hub
#   [17 = spacer]
#   18=MBRF Revenue, 19=MBRF GP, 20=MBRF GM%, 21=MBRF EBITDA, 22=MBRF EBITDA Mgn%
#   23=JBS Revenue,  24=JBS GP,  25=JBS GM%,  26=JBS EBIT,   27=JBS EBIT Mgn%
#   28=JBS EBITDA,   29=JBS EBITDA Mgn%
#   30=Tyson Sales,  31=Tyson Adj Op Inc,  32=Tyson Adj Op Mgn%

import re

def quarter_start(q: str):
    """'1Q18' → '2018-01-01'"""
    m = re.match(r"([1-4])Q(\d{2})", q)
    if not m: return None
    qn, yr = int(m.group(1)), 2000 + int(m.group(2))
    month = (qn - 1) * 3 + 1
    return date(yr, month, 1).isoformat()

def load_quarterly(wb):
    ws = wb["Summary \u2013 Quarterly"]
    rows = []
    for r in range(5, ws.max_row + 1):
        q = ws.cell(r, 1).value
        if not q or not isinstance(q, str) or not re.match(r"[1-4]Q\d{2}", str(q)):
            continue
        q = str(q)
        rows.append({
            "quarter":         q,
            "quarter_start":   quarter_start(q),
            "slaughter":       fv(ws.cell(r, 2).value),
            "ct150_steer":     fv(ws.cell(r, 3).value),
            "ct150_heifer":    fv(ws.cell(r, 4).value),
            "ct150_mixed":     fv(ws.cell(r, 5).value),
            "ct150_all":       fv(ws.cell(r, 6).value),
            "ks_steer":        fv(ws.cell(r, 7).value),
            "ks_heifer":       fv(ws.cell(r, 8).value),
            "ks_avg":          fv(ws.cell(r, 9).value),
            "ne_steer":        fv(ws.cell(r,10).value),
            "ne_heifer":       fv(ws.cell(r,11).value),
            "ne_avg":          fv(ws.cell(r,12).value),
            "choice":          fv(ws.cell(r,13).value),
            "select_":         fv(ws.cell(r,14).value),
            "drop_credit":     fv(ws.cell(r,15).value),
            "henry_hub":       fv(ws.cell(r,16).value),
            # MBRF
            "mbrf_revenue":    fv(ws.cell(r,18).value),
            "mbrf_gp":         fv(ws.cell(r,19).value),
            "mbrf_gm":         fv(ws.cell(r,20).value),
            "mbrf_ebitda":     fv(ws.cell(r,21).value),
            "mbrf_ebitda_mgn": fv(ws.cell(r,22).value),
            # JBS
            "jbs_revenue":     fv(ws.cell(r,23).value),
            "jbs_gp":          fv(ws.cell(r,24).value),
            "jbs_gm":          fv(ws.cell(r,25).value),
            "jbs_ebit":        fv(ws.cell(r,26).value),
            "jbs_ebit_mgn":    fv(ws.cell(r,27).value),
            "jbs_ebitda":      fv(ws.cell(r,28).value),
            "jbs_ebitda_mgn":  fv(ws.cell(r,29).value),
            # Tyson
            "tyson_sales":     fv(ws.cell(r,30).value),
            "tyson_adj_op_inc":fv(ws.cell(r,31).value),
            "tyson_adj_op_mgn":fv(ws.cell(r,32).value),
        })
    print(f"\nQuarterly rows read: {len(rows)}")
    batch_upsert("beef_quarterly", rows, "quarter")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"Loading workbook: {TRACKER_PATH}")
    wb = openpyxl.load_workbook(TRACKER_PATH, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    load_weekly(wb)
    load_quarterly(wb)

    print("\n✓ Historical load complete.")
