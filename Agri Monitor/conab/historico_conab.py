"""
historico_conab.py
==================
Rodar UMA VEZ para popular o banco com todo o histórico disponível.
Ao final, gera conab/conab_desde_2024.xlsx com todos os dados desde jan/2024.

Estrutura real do arquivo CONAB (descoberta via inspeção do txt):
  Colunas: ano_agricola | safra | uf | produto | id_produto |
           id_levantamento | dsc_levantamento |
           area_plantada_mil_ha | producao_mil_t | produtividade_mil_ha_mil_t

  - id_levantamento: '001' a '012' = levantamentos mensais | '099' = final consolidado
  - produtividade está em t/ha (toneladas por hectare)
  - safra: 'UNICA' | '1ª SAFRA' | '2ª SAFRA' | '3ª SAFRA'
  - uf: sigla do estado (sem linha de total Brasil — total precisa ser calculado)
"""

import requests
import pandas as pd
import sqlite3
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ──────────────────────────────────────────────────────────────
DB_PATH    = "Agri Monitor/conab/conab.db"
EXCEL_PATH = "Agri Monitor/conab/conab_desde_2024.xlsx"

URLS = {
    "graos": "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/LevantamentoGraos.txt",
    "cana":  "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/LevantamentoCana.txt",
}

# Nomes exatos da coluna 'produto' no arquivo da CONAB
# O [DEBUG] no log vai imprimir todos os nomes disponíveis — ajuste aqui se necessário
# Nomes exatos como aparecem na coluna 'produto' do arquivo CONAB.
PRODUTOS_GRAOS = [
    "SOJA",
    "MILHO",
    "ALGODAO EM PLUMA",
]

# Ano agrícola de corte para o Excel (banco guarda tudo)
ANO_CORTE_EXCEL = "2023/24"   # inclui safras >= 2023/24 (que começam em 2024)

# ── Inicializa banco ───────────────────────────────────────────────────────────
os.makedirs("Agri Monitor/conab", exist_ok=True)

conn = sqlite3.connect(DB_PATH)
conn.execute("""
    CREATE TABLE IF NOT EXISTS safra (
        ano_agricola       TEXT,
        safra              TEXT,
        uf                 TEXT,
        produto            TEXT,
        id_produto         TEXT,
        id_levantamento    INTEGER,
        dsc_levantamento   TEXT,
        area_plantada_mil_ha    REAL,
        producao_mil_t          REAL,
        produtividade_t_ha      REAL,
        updated_at         TEXT,
        PRIMARY KEY (ano_agricola, safra, uf, produto, id_levantamento)
    )
""")
conn.commit()

# ── Utilitários ────────────────────────────────────────────────────────────────
def baixa_txt(url):
    print(f"Baixando {url} ...")
    r = requests.get(url, timeout=180, verify=False)
    r.raise_for_status()
    df = pd.read_csv(io.StringIO(r.content.decode("latin1")), sep=";", dtype=str)
    df.columns = [c.strip() for c in df.columns]
    for c in df.columns:
        df[c] = df[c].str.strip()
    print(f"  -> {len(df)} linhas | colunas: {list(df.columns)}")
    return df

def parse_float(val):
    try:
        s = str(val).strip()
        # Suporta tanto ponto decimal ('263.7') quanto separador BR ('1.868,7')
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")   # '1.868,7' -> '1868.7'
        elif "," in s:
            s = s.replace(",", ".")                    # '1,7' -> '1.7'
        return float(s)
    except (ValueError, AttributeError):
        return None

def normaliza_levantamento(val):
    """Converte id_levantamento para inteiro 1-12 ou 99 (safra final).
    - 1 a 12  = levantamentos mensais
    - 99/099  = levantamento final consolidado (salvo como 99)
    - outros  = descartado
    """
    try:
        v = int(str(val).strip())
        if 1 <= v <= 12:
            return v
        if v in (99, 99):
            return 99
        return None
    except (ValueError, TypeError):
        return None

def upsert(conn, df):
    # Normaliza levantamento: 1-12 mensais + 99 safra final
    df["id_levantamento"] = df["id_levantamento"].apply(normaliza_levantamento)
    df = df[df["id_levantamento"].notna()].copy()
    df["id_levantamento"] = df["id_levantamento"].astype(int)
    if df.empty:
        return
    df.to_sql("safra", conn, if_exists="append", index=False, method="multi")
    conn.execute("""
        DELETE FROM safra WHERE rowid NOT IN (
            SELECT MAX(rowid) FROM safra
            GROUP BY ano_agricola, safra, uf, produto, id_levantamento
        )
    """)
    conn.commit()

# ── Grãos (Soja, Milho, Algodão) ──────────────────────────────────────────────
total_graos = 0
try:
    df_graos = baixa_txt(URLS["graos"])

    print(f"\n[DEBUG] Produtos disponíveis no arquivo de grãos:")
    for n in sorted(df_graos["produto"].dropna().unique()):
        print(f"  '{n}'")

    df_graos = df_graos[df_graos["produto"].isin(PRODUTOS_GRAOS)].copy()
    print(f"\n  -> {len(df_graos)} linhas após filtro de culturas")

    if df_graos.empty:
        print("  ATENÇÃO: 0 registros — ajuste PRODUTOS_GRAOS com os nomes do [DEBUG] acima.")
    else:
        df_graos["area_plantada_mil_ha"] = df_graos["area_plantada_mil_ha"].apply(parse_float)
        df_graos["producao_mil_t"]       = df_graos["producao_mil_t"].apply(parse_float)
        df_graos["produtividade_t_ha"]   = df_graos["produtividade_mil_ha_mil_t"].apply(parse_float)
        df_graos["updated_at"]           = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        df_final = df_graos[[
            "ano_agricola", "safra", "uf", "produto", "id_produto",
            "id_levantamento", "dsc_levantamento",
            "area_plantada_mil_ha", "producao_mil_t", "produtividade_t_ha", "updated_at"
        ]]

        upsert(conn, df_final)
        total_graos = len(df_final)
        print(f"  OK {total_graos} registros de grãos inseridos.")
        print("\n  Registros por produto:")
        for prod, cnt in df_final["produto"].value_counts().items():
            print(f"    {prod}: {cnt}")

except Exception as e:
    print(f"  ERRO ao processar grãos: {e}")
    raise

# ── Cana-de-Acucar ─────────────────────────────────────────────────────────────
# Estrutura real do arquivo (descoberta via DEBUG):
# ano_agricola | dsc_safra_previsao | uf | produto | id_produto |
# dsc_levantamento | id_levantamento | area_plantada_mil_ha | producao_mil_t |
# producao_acucar_mil_t | producao_etanol_anidro_mil_l |
# producao_etanol_hidratado_mil_l | producao_etanol_total_mil_l | produtcao_atr_kg_t
# Nao tem coluna de produtividade — salva produtcao_atr_kg_t no campo produtividade_t_ha
total_cana = 0
try:
    df_cana = baixa_txt(URLS["cana"])

    print(f"\n[DEBUG] Colunas do arquivo de cana: {list(df_cana.columns)}")

    col_area  = next((c for c in df_cana.columns if "AREA"         in c.upper()), None)
    col_prod  = next((c for c in df_cana.columns if c == "producao_mil_t"), None)
    col_atr   = next((c for c in df_cana.columns if "ATR"          in c.upper()), None)
    col_uf    = next((c for c in df_cana.columns if c.upper() == "UF"), None)
    col_ano   = next((c for c in df_cana.columns if "ANO"          in c.upper()), None)
    col_safra = next((c for c in df_cana.columns if "SAFRA"        in c.upper()), None)
    col_lev   = next((c for c in df_cana.columns if c == "id_levantamento"), None)
    col_dlev  = next((c for c in df_cana.columns if c == "dsc_levantamento"), None)
    col_idp   = next((c for c in df_cana.columns if c == "id_produto"), None)

    # Valida colunas obrigatorias
    faltando = [nome for nome, c in [("area", col_area), ("producao", col_prod),
                                     ("uf", col_uf), ("ano", col_ano),
                                     ("safra", col_safra), ("levantamento", col_lev)]
                if c is None]
    if faltando:
        raise ValueError(f"Colunas nao encontradas no arquivo de cana: {faltando}. "
                         f"Colunas disponiveis: {list(df_cana.columns)}")

    df_cana["ano_agricola"]         = df_cana[col_ano]
    df_cana["safra"]                = df_cana[col_safra]
    df_cana["uf"]                   = df_cana[col_uf]
    df_cana["produto"]              = "CANA-DE-ACUCAR"
    df_cana["id_produto"]           = df_cana[col_idp] if col_idp else ""
    df_cana["id_levantamento"]      = df_cana[col_lev]
    df_cana["dsc_levantamento"]     = df_cana[col_dlev] if col_dlev else ""
    df_cana["area_plantada_mil_ha"] = df_cana[col_area].apply(parse_float)
    df_cana["producao_mil_t"]       = df_cana[col_prod].apply(parse_float)
    # ATR (kg acucar / t cana) salvo no campo produtividade como melhor proxy disponivel
    df_cana["produtividade_t_ha"]   = df_cana[col_atr].apply(parse_float) if col_atr else None
    df_cana["updated_at"]           = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    df_final_c = df_cana[[
        "ano_agricola", "safra", "uf", "produto", "id_produto",
        "id_levantamento", "dsc_levantamento",
        "area_plantada_mil_ha", "producao_mil_t", "produtividade_t_ha", "updated_at"
    ]]

    upsert(conn, df_final_c)
    total_cana = len(df_final_c)
    print(f"  OK {total_cana} registros de cana inseridos.")

except Exception as e:
    print(f"  ERRO ao processar cana: {e}")
    print("  Continuando sem dados de cana...")

# ── Resumo do banco ────────────────────────────────────────────────────────────
total = conn.execute("SELECT COUNT(*) FROM safra").fetchone()[0]
print(f"\n{'='*55}")
print(f"Carga histórica concluída. Total no banco: {total} registros")
print("\nDistribuição:")
for row in conn.execute("""
    SELECT produto,
           COUNT(DISTINCT ano_agricola) as anos,
           COUNT(DISTINCT id_levantamento) as levantamentos,
           COUNT(DISTINCT uf) as ufs,
           COUNT(*) as total
    FROM safra GROUP BY produto ORDER BY produto
""").fetchall():
    print(f"  {row[0]}: {row[4]} reg | {row[1]} anos | {row[2]} levant. | {row[3]} UFs")

# ── Gera Excel com dados desde ANO_CORTE_EXCEL ─────────────────────────────────
print(f"\nGerando Excel com dados a partir de {ANO_CORTE_EXCEL}...")

df_excel = pd.read_sql(f"""
    SELECT
        ano_agricola        AS "Ano Agrícola",
        safra               AS "Safra",
        uf                  AS "UF",
        produto             AS "Produto",
        id_levantamento     AS "Levantamento",
        dsc_levantamento    AS "Desc. Levantamento",
        area_plantada_mil_ha    AS "Área Plantada (mil ha)",
        producao_mil_t          AS "Produção (mil t)",
        produtividade_t_ha      AS "Produtividade (t/ha)"
    FROM safra
    WHERE ano_agricola >= '{ANO_CORTE_EXCEL}'
    ORDER BY produto, ano_agricola, id_levantamento, uf
""", conn)

conn.close()

print(f"  -> {len(df_excel)} linhas no Excel")

# Cores por produto
CORES = {
    "SOJA":                          "E8F5E9",
    "MILHO 1ª SAFRA":                "FFFDE7",
    "MILHO 2ª SAFRA":                "FFF9C4",
    "MILHO 3ª SAFRA":                "FFF3E0",
    "MILHO TOTAL":                   "FFE0B2",
    "MILHO TOTAL (1ª+2ª SAFRAS)":    "FFE0B2",
    "MILHO TOTAL (1ª+2ª+3ª SAFRAS)": "FFE0B2",
    "ALGODAO EM PLUMA":              "F3E5F5",
    "ALGODAO TOTAL (PLUMA)":         "F3E5F5",
    "CANA-DE-ACUCAR":                "FBE9E7",
}
COR_LEV099  = "B2DFDB"   # verde água — levantamento final consolidado
COR_HEADER  = "263238"
thin = Side(style="thin", color="BDBDBD")
borda = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Aba 1: Dados completos ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Dados Completos"

ws1.append(list(df_excel.columns))
for cell in ws1[1]:
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = borda
ws1.row_dimensions[1].height = 32

for i, row in df_excel.iterrows():
    ws1.append(list(row))

for i, row_cells in enumerate(ws1.iter_rows(min_row=2, max_row=len(df_excel)+1)):
    produto = df_excel.iloc[i]["Produto"]
    lev     = str(df_excel.iloc[i]["Levantamento"])
    cor     = COR_LEV099 if lev == "099" else CORES.get(produto, "FFFFFF")
    for cell in row_cells:
        cell.fill      = PatternFill("solid", fgColor=cor)
        cell.border    = borda
        cell.alignment = Alignment(
            horizontal="center" if cell.column > 4 else "left",
            vertical="center"
        )

for i, row in enumerate(ws1.iter_rows(min_row=2, max_row=len(df_excel)+1)):
    row[6].number_format = "#,##0.0"
    row[7].number_format = "#,##0.0"
    row[8].number_format = "#,##0.000"

larguras = [13, 10, 5, 30, 12, 16, 22, 18, 20]
for i, l in enumerate(larguras, 1):
    ws1.column_dimensions[get_column_letter(i)].width = l
ws1.freeze_panes = "A2"

# ── Aba 2: Total Brasil por produto/ano/levantamento ──────────────────────────
ws2 = wb.create_sheet("Total Brasil")

df_br = df_excel[df_excel["Área Plantada (mil ha)"] > 0].copy()
df_br["prod_x_area"] = df_br["Produtividade (t/ha)"] * df_br["Área Plantada (mil ha)"]

df_total = df_br.groupby(
    ["Produto", "Ano Agrícola", "Safra", "Levantamento", "Desc. Levantamento"],
    sort=False
).agg(
    area_total    = ("Área Plantada (mil ha)", "sum"),
    producao_total= ("Produção (mil t)", "sum"),
    prod_x_area   = ("prod_x_area", "sum"),
    area_peso     = ("Área Plantada (mil ha)", "sum"),
).reset_index()

df_total["Produtividade Média (t/ha)"] = (df_total["prod_x_area"] / df_total["area_peso"]).round(3)
df_total = df_total.sort_values(["Produto", "Ano Agrícola", "Levantamento"]).reset_index(drop=True)
df_total = df_total.rename(columns={
    "area_total":     "Área Total Brasil (mil ha)",
    "producao_total": "Produção Total Brasil (mil t)"
})

cols2 = ["Produto","Ano Agrícola","Safra","Levantamento","Desc. Levantamento",
         "Área Total Brasil (mil ha)","Produção Total Brasil (mil t)","Produtividade Média (t/ha)"]

ws2.append(cols2)
for cell in ws2[1]:
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = borda
ws2.row_dimensions[1].height = 32

for i, row in df_total.iterrows():
    ws2.append([row[c] for c in cols2])

for i, row_cells in enumerate(ws2.iter_rows(min_row=2, max_row=len(df_total)+1)):
    produto = df_total.iloc[i]["Produto"]
    lev     = str(df_total.iloc[i]["Levantamento"])
    cor     = COR_LEV099 if lev == "099" else CORES.get(produto, "FFFFFF")
    for cell in row_cells:
        cell.fill      = PatternFill("solid", fgColor=cor)
        cell.border    = borda
        cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left", vertical="center")

for row in ws2.iter_rows(min_row=2, max_row=len(df_total)+1):
    row[5].number_format = "#,##0.0"
    row[6].number_format = "#,##0.0"
    row[7].number_format = "#,##0.000"

for i, l in enumerate([30,13,10,12,16,26,26,24], 1):
    ws2.column_dimensions[get_column_letter(i)].width = l
ws2.freeze_panes = "A2"

# Legenda
ws2.append([])
ws2.append(["Legenda:"])
ws2.append(["Lev. 001-012 = levantamentos mensais do ano agrícola"])
ws2.append(["Lev. 099 (verde água) = levantamento final consolidado da safra"])
ws2.append(["Produtividade = t/ha ponderada pela área de cada estado"])
ws2.append([f"Fonte: CONAB | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])

wb.save(EXCEL_PATH)
print(f"  Excel salvo em: {EXCEL_PATH}")
print(f"\nConcluído.")
