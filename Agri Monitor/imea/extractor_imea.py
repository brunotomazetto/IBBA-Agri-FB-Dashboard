import hashlib
import io
import logging
import os
import re
import sqlite3
import time
from datetime import datetime

import openpyxl
import requests

# ── Configurações ──────────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(__file__), "imea.db")

# API pública IMEA (custos — Excel mensal)
API_ARQUIVO = "https://api1.imea.com.br/api/arquivo"
TIPO_CUSTO  = "696277432068079616"

# CONAB (preços mensais por UF — gratuito, sem login)
URL_PRECO_CONAB = "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/PrecosMensalUF.txt"

CULTURAS_IMEA = {
    "SOJA":    "4",
    "MILHO":   "3",
    "ALGODAO": "1",
}

# Produtos CONAB por cultura (nome exato no arquivo, nível MT)
PRODUTOS_CONAB = {
    "SOJA":    ["SOJA EM GRAO", "SOJA"],
    "MILHO":   ["MILHO EM GRAO", "MILHO"],
    "ALGODAO": ["ALGODAO EM PLUMA", "ALGODAO"],
}

HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": "https://imea.com.br/"}

MESES_PT = {
    "janeiro":1,"fevereiro":2,"março":3,"abril":4,"maio":5,"junho":6,
    "julho":7,"agosto":8,"setembro":9,"outubro":10,"novembro":11,"dezembro":12,
    "jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
    "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12,
}

NOME_PARA_INDICADOR = {
    "a. custeio":                    "Custeio",
    "coe (a + b + ... + f + g)":    "Custo Operacional Efetivo",
    "cot (coe + h + i)":            "Custo Operacional Total",
    "ct (cot + j)":                 "Custo Total",
    "1. sementes":                   "Sementes",
    "2. fertilizantes e corretivos": "Fertilizantes e Corretivos",
    "3. defensivos":                 "Defensivos",
    "4. operações mecanizadas":      "Operações Mecanizadas",
    "5. serviços terceirizados":     "Serviços Terceirizados",
    "6. mão de obra":                "Mão de Obra",
    "b. manutenção":                 "Manutenção",
    "c. impostos e taxas":           "Impostos e Taxas",
    "d. financeiras":                "Financeiras",
    "e. pós-produção":               "Pós-Produção",
    "f. outros custos":              "Outros Custos",
    "g. arrendamento":               "Arrendamento",
    "h. depreciações":               "Depreciações",
    "i. mão-de-obra familiar":       "Mão-de-obra Familiar",
    "j. custo de oportunidade":      "Custo de Oportunidade",
    "custeio":                       "Custeio",
    "operações mecanizadas":         "Operações Mecanizadas",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── Inicializa banco ───────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
conn = sqlite3.connect(DB_PATH)

conn.executescript("""
    CREATE TABLE IF NOT EXISTS historico (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        cultura          TEXT    NOT NULL,
        cadeia_id        TEXT,
        indicador_id     TEXT,
        indicador_nome   TEXT,
        safra            TEXT,
        safra_id         TEXT,
        safra_tipo       TEXT,
        data_referencia  TEXT,
        ano              INTEGER,
        mes              INTEGER,
        valor            REAL,
        unidade          TEXT,
        estado           TEXT,
        grupo            TEXT,
        updated_at       TEXT,
        UNIQUE(indicador_id, safra_id, data_referencia, estado)
    );
    CREATE TABLE IF NOT EXISTS arquivos_processados (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        imea_id     TEXT    UNIQUE,
        cultura     TEXT,
        nome        TEXT,
        hash_md5    TEXT    UNIQUE,
        coletado_em TEXT
    );
    CREATE TABLE IF NOT EXISTS preco_conab (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        cultura          TEXT    NOT NULL,
        produto_conab    TEXT,
        uf               TEXT,
        ano              INTEGER,
        mes              INTEGER,
        data_referencia  TEXT,
        nivel_comercializacao TEXT,
        valor_kg         REAL,
        updated_at       TEXT,
        UNIQUE(produto_conab, uf, ano, mes, nivel_comercializacao)
    );
    CREATE INDEX IF NOT EXISTS idx_hist_cultura  ON historico(cultura);
    CREATE INDEX IF NOT EXISTS idx_hist_grupo    ON historico(grupo);
    CREATE INDEX IF NOT EXISTS idx_preco_cultura ON preco_conab(cultura);
""")
conn.commit()

# ── Utilitários ────────────────────────────────────────────────────────────────
def md5(content: bytes) -> str:
    return hashlib.md5(content).hexdigest()

def hash_existe(h: str) -> bool:
    return conn.execute(
        "SELECT 1 FROM arquivos_processados WHERE hash_md5=?", (h,)
    ).fetchone() is not None

def normaliza(nome: str) -> str:
    return re.sub(r"\*+$", "", str(nome).strip()).strip().lower()

def extrai_mes(nome: str) -> int | None:
    n = normaliza(nome)
    for k, v in MESES_PT.items():
        if k in n:
            return v
    return None

def parse_float(val) -> float | None:
    try:
        return float(str(val).replace(",", "."))
    except (TypeError, ValueError):
        return None

# ── PARTE 1: Custos + Produtividade via Excel IMEA ─────────────────────────────
def parse_excel(content: bytes, cultura: str) -> list[dict]:
    wb   = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "indice" or not sheet_name.upper().endswith("_MT"):
            continue

        ws = wb[sheet_name]
        safra_row = ano_row = mes_row = None

        for r in range(1, 15):
            c0 = normaliza(ws.cell(r, 1).value or "")
            if c0 == "safra":               safra_row = r
            elif c0 == "ano":               ano_row   = r
            elif c0 in ("mês", "mes"):      mes_row   = r

        if not mes_row:
            continue

        colunas = {}
        for c in range(2, ws.max_column + 1):
            safra    = str(ws.cell(safra_row, c).value or "").strip() if safra_row else ""
            ano      = str(ws.cell(ano_row,   c).value or "").strip() if ano_row   else ""
            mes_nome = str(ws.cell(mes_row,   c).value or "").strip()
            mes_num  = extrai_mes(mes_nome)
            if not mes_num or not safra:
                continue
            try:
                ano_int = int(float(ano))
            except:
                continue
            colunas[c] = {"safra": safra, "ano": ano_int, "mes": mes_num,
                          "data_ref": f"{ano_int}-{mes_num:02d}-15"}

        for r in range(mes_row + 1, ws.max_row + 1):
            nome_raw = str(ws.cell(r, 1).value or "").strip()
            if not nome_raw:
                continue

            nome_lower = normaliza(nome_raw)

            # Produtividade Modal — capturar separado como PRODUTIVIDADE
            if nome_lower.startswith("produtividade modal"):
                for c, ci in colunas.items():
                    val = parse_float(ws.cell(r, c).value)
                    if val is None:
                        continue
                    rows.append({
                        "cultura":        cultura,
                        "cadeia_id":      CULTURAS_IMEA[cultura],
                        "indicador_id":   None,
                        "indicador_nome": "Produtividade Modal",
                        "safra":          ci["safra"],
                        "safra_id":       None,
                        "safra_tipo":     None,
                        "data_referencia": ci["data_ref"],
                        "ano":            ci["ano"],
                        "mes":            ci["mes"],
                        "valor":          val,
                        "unidade":        "sc/ha",
                        "estado":         "MT",
                        "grupo":          "PRODUTIVIDADE",
                        "updated_at":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                continue

            # Ignorar rodapés
            if nome_lower.startswith(("fonte","nota","unidade","*","dólar","dolar","produtividade")):
                continue

            # Resolver nome do indicador (custos)
            ind_nome = NOME_PARA_INDICADOR.get(nome_lower)
            if not ind_nome:
                nc = re.sub(r"^[\w]\.\s+", "", nome_raw).strip()
                nc = re.sub(r"\*+$", "", nc).strip()
                nc = re.sub(r"\s*\(.*?\)\s*$", "", nc).strip()
                ind_nome = nc

            for c, ci in colunas.items():
                val = parse_float(ws.cell(r, c).value)
                if val is None:
                    continue
                rows.append({
                    "cultura":        cultura,
                    "cadeia_id":      CULTURAS_IMEA[cultura],
                    "indicador_id":   None,
                    "indicador_nome": ind_nome,
                    "safra":          ci["safra"],
                    "safra_id":       None,
                    "safra_tipo":     None,
                    "data_referencia": ci["data_ref"],
                    "ano":            ci["ano"],
                    "mes":            ci["mes"],
                    "valor":          val,
                    "unidade":        "R$/ha",
                    "estado":         "MT",
                    "grupo":          "CUSTO",
                    "updated_at":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
    return rows


def upsert_historico(rows: list[dict]) -> int:
    if not rows:
        return 0
    conn.executemany(
        """INSERT OR IGNORE INTO historico
           (cultura, cadeia_id, indicador_id, indicador_nome, safra, safra_id,
            safra_tipo, data_referencia, ano, mes, valor, unidade, estado,
            grupo, updated_at)
           VALUES (:cultura,:cadeia_id,:indicador_id,:indicador_nome,:safra,:safra_id,
                   :safra_tipo,:data_referencia,:ano,:mes,:valor,:unidade,:estado,
                   :grupo,:updated_at)""",
        rows,
    )
    conn.commit()
    return len(rows)


def lista_relatorios(cadeia_id: str) -> list[dict]:
    todos, page = [], 1
    while True:
        url = f"{API_ARQUIVO}?cadeia={cadeia_id}&tipo={TIPO_CUSTO}&page={page}&pageSize=100&nome=&sort=1"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            log.error(f"  Erro API: {e}")
            break
        todos.extend(data.get("Result", []))
        if page >= data.get("PageCount", 1):
            break
        page += 1
        time.sleep(0.3)
    return todos


def processa_custos():
    log.info("\n=== CUSTOS + PRODUTIVIDADE (Excel IMEA) ===")
    for cultura, cadeia_id in CULTURAS_IMEA.items():
        log.info(f"\n[{cultura}]")
        relatorios = lista_relatorios(cadeia_id)
        log.info(f"  {len(relatorios)} relatório(s)")

        for rel in relatorios:
            imea_id = rel.get("Id", "")
            nome    = rel.get("Nome", "")
            url_s3  = rel.get("Path", "")
            if not url_s3:
                continue
            log.info(f"  → {nome}")
            try:
                content = requests.get(url_s3, headers=HEADERS, timeout=60).content
            except Exception as e:
                log.error(f"    Download: {e}")
                continue

            h = md5(content)
            if hash_existe(h):
                log.info(f"    Já processado — pulando.")
                continue

            try:
                rows = parse_excel(content, cultura)
            except Exception as e:
                log.error(f"    Parse: {e}", exc_info=True)
                continue

            custos = [r for r in rows if r["grupo"] == "CUSTO"]
            produt = [r for r in rows if r["grupo"] == "PRODUTIVIDADE"]
            ins_c  = upsert_historico(custos)
            ins_p  = upsert_historico(produt)

            conn.execute(
                "INSERT OR IGNORE INTO arquivos_processados (imea_id,cultura,nome,hash_md5,coletado_em) VALUES (?,?,?,?,?)",
                (imea_id, cultura, nome, h, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            conn.commit()
            log.info(f"    {ins_c} custos + {ins_p} produtividades inseridos")
            time.sleep(0.5)

        if not relatorios:
            log.info(f"  Nenhum relatório novo.")


# ── PARTE 2: Preços mensais via CONAB ─────────────────────────────────────────
def processa_precos_conab():
    log.info("\n=== PREÇOS MENSAIS (CONAB) ===")
    try:
        resp = requests.get(URL_PRECO_CONAB, timeout=120, verify=False)
        resp.raise_for_status()
        content = resp.content.decode("latin1")
    except Exception as e:
        log.error(f"  Erro download CONAB: {e}")
        return

    linhas = content.splitlines()
    if not linhas:
        return

    cabecalho = [c.strip().lower() for c in linhas[0].split(";")]
    log.info(f"  Colunas: {cabecalho}")

    # Mapear todos os nomes de produto CONAB que queremos
    produtos_alvo = {}
    for cultura, nomes in PRODUTOS_CONAB.items():
        for nome in nomes:
            produtos_alvo[nome.upper().strip()] = cultura

    registros = []
    for linha in linhas[1:]:
        cols = [c.strip() for c in linha.split(";")]
        if len(cols) < len(cabecalho):
            continue

        row = dict(zip(cabecalho, cols))
        produto = row.get("produto", "").strip().upper()
        uf      = row.get("uf", "").strip().upper()

        # Filtrar: só MT e produtos alvo
        if uf != "MT":
            continue

        cultura = None
        for nome_alvo, cult in produtos_alvo.items():
            if nome_alvo in produto:
                cultura = cult
                break
        if not cultura:
            continue

        try:
            ano = int(row.get("ano", 0))
            mes = int(row.get("mes", 0))
            val = float(row.get("valor_produto_kg", "0").replace(",", "."))
        except (ValueError, TypeError):
            continue

        if ano < 2020 or mes < 1 or mes > 12 or val <= 0:
            continue

        nivel = row.get("dsc_nivel_comercializacao", "").strip()
        registros.append((
            cultura,
            produto,
            uf,
            ano,
            mes,
            f"{ano}-{mes:02d}-15",
            nivel,
            val,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))

    if registros:
        conn.executemany(
            """INSERT OR REPLACE INTO preco_conab
               (cultura, produto_conab, uf, ano, mes, data_referencia,
                nivel_comercializacao, valor_kg, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            registros,
        )
        conn.commit()
        log.info(f"  {len(registros)} preços inseridos/atualizados (MT)")

        # Resumo por cultura
        for row in conn.execute("""
            SELECT cultura, COUNT(*) as n, MIN(data_referencia), MAX(data_referencia)
            FROM preco_conab GROUP BY cultura ORDER BY cultura
        """):
            log.info(f"  {row[0]}: {row[1]} registros | {row[2]} → {row[3]}")
    else:
        log.info("  Nenhum preço encontrado para MT")


# ── PARTE 3: Produtividade via CONAB safra ────────────────────────────────────
URL_LEVANTAMENTO_CONAB = "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/LevantamentoGraos.txt"

# Levantamento → (mês, delta_ano)
# Soja/Algodão: safra out-set  → Lev1=out(ano1), Lev4=jan(ano2)...
# Milho 2ª safra: safra jan-dez → Lev1=jan(ano2), Lev12=dez(ano2)
LEV_SOJA_ALGODAO = {
    1:(10,0), 2:(11,0), 3:(12,0), 4:(1,1),  5:(2,1),  6:(3,1),
    7:(4,1),  8:(5,1),  9:(6,1),  10:(7,1), 11:(8,1), 12:(9,1),
}
LEV_MILHO = {
    1:(1,1), 2:(2,1),  3:(3,1),  4:(4,1),  5:(5,1),  6:(6,1),
    7:(7,1), 8:(8,1),  9:(9,1),  10:(10,1),11:(11,1),12:(12,1),
}
CONAB_PRODUTO_CULTURA = {
    "SOJA":             ("SOJA",    LEV_SOJA_ALGODAO),
    "MILHO":            ("MILHO",   LEV_MILHO),
    "ALGODAO EM PLUMA": ("ALGODAO", LEV_SOJA_ALGODAO),
}


def processa_produtividade_conab():
    log.info("\n=== PRODUTIVIDADE (CONAB Levantamento Grãos) ===")
    try:
        resp = requests.get(URL_LEVANTAMENTO_CONAB, timeout=120, verify=False)
        resp.raise_for_status()
        content = resp.content.decode("latin1")
    except Exception as e:
        log.error(f"  Erro download CONAB levantamento: {e}")
        return

    linhas = content.splitlines()
    if not linhas:
        return

    cabecalho = [c.strip().lower() for c in linhas[0].split(";")]
    registros = []

    for linha in linhas[1:]:
        cols = [c.strip() for c in linha.split(";")]
        if len(cols) < len(cabecalho):
            continue
        row = dict(zip(cabecalho, cols))

        if row.get("uf", "").strip().upper() != "MT":
            continue

        produto = row.get("produto", "").strip().upper()
        if produto not in CONAB_PRODUTO_CULTURA:
            continue

        try:
            lev = int(row.get("id_levantamento", 0))
        except:
            continue
        if lev not in range(1, 13):
            continue

        cultura, lev_map = CONAB_PRODUTO_CULTURA[produto]
        mes, delta = lev_map[lev]

        ano_agr = row.get("ano_agricola", "").strip()
        try:
            ano1 = int(ano_agr[:4])
            ano  = ano1 + delta
        except:
            continue

        try:
            prod_t_ha = float(row.get("produtividade_mil_ha_mil_t", "0").replace(",", "."))
        except:
            continue
        if prod_t_ha <= 0:
            continue

        prod_sc_ha = round(prod_t_ha * 1000 / 60, 4)
        data_ref   = f"{ano}-{mes:02d}-15"

        registros.append((
            cultura, None,
            f"conab_prod_{cultura.lower()}_{lev}",
            "Produtividade CONAB",
            ano_agr, None, None,
            data_ref, ano, mes,
            prod_sc_ha, "sc/ha", "MT",
            "PRODUTIVIDADE",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))

    if registros:
        conn.executemany("""
            INSERT OR REPLACE INTO historico
            (cultura, cadeia_id, indicador_id, indicador_nome, safra, safra_id,
             safra_tipo, data_referencia, ano, mes, valor, unidade, estado,
             grupo, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, registros)
        conn.commit()
        log.info(f"  {len(registros)} registros inseridos/atualizados")
        for row in conn.execute("""
            SELECT cultura, COUNT(*), MIN(data_referencia), MAX(data_referencia)
            FROM historico WHERE grupo='PRODUTIVIDADE' AND indicador_nome='Produtividade CONAB'
            GROUP BY cultura ORDER BY cultura
        """):
            log.info(f"  {row[0]}: {row[1]} | {row[2]} → {row[3]}")
    else:
        log.info("  Nenhum dado encontrado para MT")


# ── Main ───────────────────────────────────────────────────────────────────────
processa_custos()
processa_precos_conab()
processa_produtividade_conab()

conn.close()
print("\nConcluído.")
